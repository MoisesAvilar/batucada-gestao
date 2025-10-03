from decimal import Decimal

from django.apps import apps
from django.http import JsonResponse
from collections import defaultdict
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import (
    Transaction,
    Despesa,
    Receita,
    DespesaRecorrente,
    ReceitaRecorrente,
    Category,
)
from store.models import Produto
from .forms import (
    TransactionForm,
    CategoryForm,
    DespesaForm,
    DespesaRecorrenteForm,
    ReceitaRecorrenteForm,
    MensalidadeReceitaForm,
    VendaReceitaForm,
)

from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from django.db.models import Sum, Count
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models.functions import TruncMonth
from datetime import date, timedelta, datetime
from django.db.models import Q
from scheduler.models import Aula, CustomUser, Modalidade, PresencaAluno
from django.utils.timezone import now
from functools import wraps
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from scheduler.models import Aluno

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.cell.rich_text import TextBlock, Text
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa


def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if getattr(request.user, "tipo", None) == "admin":
            return view_func(request, *args, **kwargs)
        return redirect("scheduler:dashboard")

    return wrapper


def add_months(start_date, months):
    month = start_date.month - 1 + months
    year = start_date.year + month // 12
    month = month % 12 + 1
    day = min(
        start_date.day, [0, 31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month]
    )
    return date(year, month, day)


@admin_required
def transaction_list_view(request):
    MESES_PT = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
        7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
    }
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.warning(request, "Selecione uma Unidade de Negócio.")
        return redirect("scheduler:dashboard")

    today = now().date()
    ontem = today - timedelta(days=1)
    start_date = (
        date.fromisoformat(request.GET.get("start_date"))
        if request.GET.get("start_date")
        else today.replace(day=1)
    )
    end_date = (
        date.fromisoformat(request.GET.get("end_date"))
        if request.GET.get("end_date")
        else (today.replace(day=28) + timedelta(days=4)).replace(day=1)
        - timedelta(days=1)
    )

    # ... (o restante da lógica de KPIs e contas vencidas permanece igual) ...
    total_a_pagar = Despesa.objects.filter(unidade_negocio_id=unidade_ativa_id, status="a_pagar", data_competencia__range=[start_date, end_date]).aggregate(total=Sum("valor"))["total"] or 0
    total_a_receber = Receita.objects.filter(unidade_negocio_id=unidade_ativa_id, status="a_receber", data_competencia__range=[start_date, end_date]).aggregate(total=Sum("valor"))["total"] or 0
    contas_vencidas = Despesa.objects.filter(unidade_negocio_id=unidade_ativa_id, status='a_pagar', data_competencia__lt=today).aggregate(count=Count('id'), total=Sum('valor'))
    receitas_atrasadas = Receita.objects.filter(unidade_negocio_id=unidade_ativa_id, status='a_receber', data_competencia__lt=today).aggregate(count=Count('id'), total=Sum('valor'))
    transactions = Transaction.objects.filter(unidade_negocio_id=unidade_ativa_id, transaction_date__range=[start_date, end_date]).select_related("category")
    no_data = not transactions.exists()
    # ==============================================================================
    # --- FLUXO DE CAIXA REALISTA (COM MESES EM PORTUGUÊS) ---
    # ==============================================================================
    def get_monthly_data(transactions, tipo):
        qs = transactions.filter(category__type=tipo)
        monthly = qs.annotate(month=TruncMonth("transaction_date")) \
                      .values("month") \
                      .annotate(total=Sum("amount")) \
                      .order_by("month")
        # ★ ALTERAÇÃO 1: Usar o objeto 'date' como chave, em vez de uma string formatada
        month_map = {d["month"]: float(d["total"]) for d in monthly}
        return month_map

    income_map = get_monthly_data(transactions, "income")
    expense_map = get_monthly_data(transactions, "expense")

    # Obtém um conjunto de todas as datas (objetos date)
    months_set = set(income_map.keys()).union(expense_map.keys())
    # Ordena as datas cronologicamente
    flow_chart_dates = sorted(list(months_set))

    # ★ ALTERAÇÃO 2: Cria as legendas em português usando o dicionário MESES_PT
    flow_chart_labels = [f"{MESES_PT[d.month]}/{d.strftime('%y')}" for d in flow_chart_dates]

    # ★ ALTERAÇÃO 3: Busca os dados usando as chaves de data ordenadas
    flow_chart_income_data = [income_map.get(m, 0) for m in flow_chart_dates]
    flow_chart_expense_data = [expense_map.get(m, 0) for m in flow_chart_dates]
    
    # ==============================================================================
    # --- FIM DAS ALTERAÇÕES NO FLUXO DE CAIXA ---
    # ==============================================================================

    def rolling_average(data, window=3):
        # ... (função rolling_average sem alterações) ...
        result = []
        for i in range(len(data)):
            slice_data = data[max(0, i-window+1):i+1]
            result.append(sum(slice_data)/len(slice_data))
        return result

    flow_chart_income_avg = rolling_average(flow_chart_income_data)
    flow_chart_expense_avg = rolling_average(flow_chart_expense_data)

    total_income = sum(flow_chart_income_data)
    # total_expenses terá um valor negativo (ex: -21941.84)
    total_expenses = sum(flow_chart_expense_data) 
    
    balance = total_income + total_expenses
    
    expenses_by_category = transactions.filter(category__type="expense").values("category__name").annotate(total=Sum("amount")).order_by("-total")
    chart_labels = [item["category__name"] for item in expenses_by_category]
    chart_data = [float(item["total"]) for item in expenses_by_category]

    income_by_category = transactions.filter(category__type="income").values("category__name").annotate(total=Sum("amount")).order_by("-total")
    income_chart_labels = [item["category__name"] for item in income_by_category]
    income_chart_data = [float(item["total"]) for item in income_by_category]
    
    paginator = Paginator(transactions.order_by('-transaction_date'), 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    form = TransactionForm(initial={"unidade_negocio": unidade_ativa_id})
    if request.method == "POST":
        form = TransactionForm(request.POST)
        if form.is_valid():
            t = form.save(commit=False)
            t.created_by = request.user
            t.unidade_negocio_id = unidade_ativa_id
            t.save()
            messages.success(request, "Lançamento adicionado com sucesso!")
            return redirect(request.get_full_path())

    return render(request, "finances/transaction_list.html", {
        "page_obj": page_obj,
        "form": form,
        "total_income": total_income,
        "total_expenses": abs(total_expenses),
        "balance": balance,
        "start_date": start_date,
        "end_date": end_date,
        "data_ontem": ontem,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "income_chart_labels": income_chart_labels,
        "income_chart_data": income_chart_data,
        "total_a_pagar": total_a_pagar,
        "total_a_receber": total_a_receber,
        'contas_vencidas_count': contas_vencidas['count'] or 0,
        'contas_vencidas_total': contas_vencidas['total'] or 0,
        'receitas_atrasadas_count': receitas_atrasadas['count'] or 0,
        'receitas_atrasadas_total': receitas_atrasadas['total'] or 0,
        'flow_chart_labels': flow_chart_labels,
        'flow_chart_income_data': flow_chart_income_data,
        'flow_chart_expense_data': flow_chart_expense_data,
        'flow_chart_income_avg': flow_chart_income_avg,
        'flow_chart_expense_avg': flow_chart_expense_avg,
        'top_produtos_vendidos': Produto.objects.filter(unidade_negocio_id=unidade_ativa_id, receitas__data_competencia__range=[start_date, end_date]).annotate(total_vendido=Sum('receitas__valor')).order_by('-total_vendido').filter(total_vendido__gt=0)[:5],
        'top_despesas': transactions.filter(category__type='expense').order_by('amount')[:5],
        'no_data': no_data,
    })


@admin_required
def add_category_ajax(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            return JsonResponse(
                {"status": "success", "id": category.id, "name": category.name}
            )
        else:
            return JsonResponse({"status": "error", "errors": form.errors})
    return JsonResponse({"status": "error", "message": "Invalid request method"})


@admin_required
@require_POST
def add_mensalidade(request):
    """
    View dedicada para adicionar MENSALIDADE, seguindo os padrões do projeto.
    """
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.error(
            request, "Sessão expirada ou nenhuma unidade de negócio selecionada."
        )
        return redirect("finances:receita_list")

    form = MensalidadeReceitaForm(request.POST)
    if form.is_valid():
        try:
            with transaction.atomic():
                receita = form.save(commit=False)
                receita.unidade_negocio_id = unidade_ativa_id  # Padrão do projeto

                # Se já foi recebido, cria a transação e atualiza o status
                if receita.data_recebimento:
                    receita.status = "recebido"
                    transacao = Transaction.objects.create(
                        unidade_negocio_id=unidade_ativa_id,
                        description=f"Recebimento: {receita.descricao}",
                        amount=receita.valor,
                        category=receita.categoria,
                        transaction_date=receita.data_recebimento,
                        student=receita.aluno,
                        created_by=request.user,
                    )
                    receita.transacao = transacao

                receita.save()
                messages.success(request, "Mensalidade adicionada com sucesso!")

        except Exception as e:
            messages.error(request, f"Ocorreu um erro inesperado: {e}")
    else:
        erros = ". ".join(
            [f"{field}: {error[0]}" for field, error in form.errors.items()]
        )
        messages.error(request, f"Erro ao adicionar mensalidade: {erros}")

    return redirect("finances:receita_list")


@admin_required
@require_POST
def add_venda(request):
    """
    View dedicada para adicionar VENDA, seguindo os padrões do projeto.
    """
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.error(
            request, "Sessão expirada ou nenhuma unidade de negócio selecionada."
        )
        return redirect("finances:receita_list")

    form = VendaReceitaForm(request.POST)
    if form.is_valid():
        try:
            with transaction.atomic():
                receita = form.save(commit=False)
                receita.unidade_negocio_id = unidade_ativa_id

                produto = receita.produto
                quantidade_vendida = receita.quantidade

                if produto.quantidade_em_estoque < quantidade_vendida:
                    messages.error(
                        request, f"Estoque insuficiente para '{produto.nome}'."
                    )
                    return redirect("finances:receita_list")

                produto.quantidade_em_estoque -= quantidade_vendida
                produto.save()

                # Se já foi recebido, cria a transação e atualiza o status
                if receita.data_recebimento:
                    receita.status = "recebido"
                    transacao = Transaction.objects.create(
                        unidade_negocio_id=unidade_ativa_id,
                        description=f"Recebimento: {receita.descricao}",
                        amount=receita.valor,
                        category=receita.categoria,
                        transaction_date=receita.data_recebimento,
                        student=receita.aluno,
                        created_by=request.user,
                    )
                    receita.transacao = transacao

                receita.save()
                messages.success(
                    request,
                    f'Venda de "{produto.nome}" registrada e estoque atualizado!',
                )

        except Exception as e:
            messages.error(request, f"Ocorreu um erro inesperado: {e}")
    else:
        erros = ". ".join(
            [f"{field}: {error[0]}" for field, error in form.errors.items()]
        )
        messages.error(request, f"Erro ao registrar venda: {erros}")

    return redirect("finances:receita_list")


@admin_required
def despesa_list_view(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.warning(request, "Selecione uma Unidade de Negócio.")
        return redirect("scheduler:dashboard")

    # --- Processamento do Formulário (POST) ---
    # Colocamos a lógica de POST no início para que, após salvar,
    # a página seja recarregada já com os dados atualizados.
    if request.method == "POST":
        form = DespesaForm(request.POST)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.unidade_negocio_id = unidade_ativa_id

            if despesa.data_pagamento:
                despesa.status = "pago"
                # Cria a transação associada se a despesa já foi paga
                transacao = Transaction.objects.create(
                    unidade_negocio_id=unidade_ativa_id,
                    description=f"Pagamento: {despesa.descricao}",
                    amount=despesa.valor,
                    category=despesa.categoria,
                    transaction_date=despesa.data_pagamento,
                    professor=despesa.professor,
                    created_by=request.user,
                )
                despesa.transacao = transacao

            despesa.save()
            messages.success(request, "Despesa registrada com sucesso!")
            return redirect("finances:despesa_list")
        # Se o form for inválido, ele será renderizado com os erros abaixo
    else:
        form = DespesaForm()

    # --- Lógica de Listagem e Filtros (GET) ---

    # 1. Capturar parâmetros de filtro
    descricao = request.GET.get('descricao', '')
    data_inicial = request.GET.get('data_inicial', '')
    data_final = request.GET.get('data_final', '')
    professor_id = request.GET.get('professor', '')
    categoria_id = request.GET.get('categoria', '')
    status = request.GET.get('status', '')

    # --- Início do Cálculo dos KPIs ---
    # Esta queryset é apenas para os KPIs, refletindo os totais da unidade de negócio
    base_qs_kpi = Despesa.objects.filter(unidade_negocio_id=unidade_ativa_id)
    
    total_a_pagar = base_qs_kpi.filter(status='a_pagar').aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    pago_qs = base_qs_kpi.filter(status='pago')
    if data_inicial: 
        pago_qs = pago_qs.filter(data_pagamento__gte=data_inicial)
    if data_final: 
        pago_qs = pago_qs.filter(data_pagamento__lte=data_final)
    total_pago = pago_qs.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    
    total_recorrentes = DespesaRecorrente.objects.filter(unidade_negocio_id=unidade_ativa_id, ativa=True).count()
    
    today = now().date()
    data_limite_vencimento = today + timedelta(days=5)
    total_a_vencer = base_qs_kpi.filter(
        status='a_pagar',
        data_competencia__gte=today,
        data_competencia__lte=data_limite_vencimento
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    # --- Fim do Cálculo dos KPIs ---

    # 2. Queryset principal para a lista, com filtros aplicados
    despesas_list = Despesa.objects.filter(unidade_negocio_id=unidade_ativa_id).select_related("categoria", "professor")

    if descricao:
        despesas_list = despesas_list.filter(descricao__icontains=descricao)
    if data_inicial:
        despesas_list = despesas_list.filter(data_competencia__gte=data_inicial)
    if data_final:
        despesas_list = despesas_list.filter(data_competencia__lte=data_final)
    if professor_id:
        despesas_list = despesas_list.filter(professor_id=professor_id)
    if categoria_id:
        despesas_list = despesas_list.filter(categoria_id=categoria_id)
    if status:
        despesas_list = despesas_list.filter(status=status)

    orderby = request.GET.get('orderby', '-data_competencia')
    allowed_orderby_fields = ['descricao', '-descricao', 'valor', '-valor', 'data_competencia', '-data_competencia']
    if orderby not in allowed_orderby_fields:
        orderby = '-data_competencia'

    despesas_list = despesas_list.order_by(orderby)
    titulo = "Saídas"

    filtro_ativo = request.GET.get("filtro")
    if filtro_ativo == "a_vencer":
        today = now().date()
        data_limite_vencimento = today + timedelta(days=5)
        despesas_list = despesas_list.filter(
            status="a_pagar",
            data_competencia__gte=today,
            data_competencia__lte=data_limite_vencimento,
        ).order_by("data_competencia")
        titulo = "Despesas a Vencer"

    # Paginação
    paginator = Paginator(despesas_list, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    # Adiciona flag 'is_vencida' para o template
    for despesa in page_obj:
        despesa.is_vencida = despesa.data_competencia < today and despesa.status == 'a_pagar'

    # 3. Montar o contexto para o template
    context = {
        "form": form,
        "page_obj": page_obj,
        "titulo": titulo,
        "orderby": orderby,
        "filtro_ativo": filtro_ativo,
        "professores": CustomUser.objects.filter(tipo__in=['admin', 'professor']),
        "categorias": Category.objects.filter(type='expense'),
        "status_choices": Despesa.STATUS_CHOICES,
        "filtros_aplicados": {
            'descricao': descricao,
            'data_inicial': data_inicial,
            'data_final': data_final,
            'professor': professor_id,
            'categoria': categoria_id,
            'status': status,
        },
        "today": today,
        # Adicionando os KPIs ao contexto
        "total_a_pagar": total_a_pagar,
        "total_pago": total_pago,
        "total_recorrentes": total_recorrentes,
        "total_a_vencer": total_a_vencer,
    }

    return render(request, "finances/despesa_list.html", context)


@admin_required
def baixar_despesa_view(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    if request.method == "POST":
        data_pagamento_str = request.POST.get("data_pagamento")
        juros_multa_str = request.POST.get("juros_multa", "0")

        if data_pagamento_str:
            data_pagamento = date.fromisoformat(data_pagamento_str)
            juros_multa = Decimal(juros_multa_str or 0)

            try:
                with transaction.atomic():
                    # Cria transação principal (sem alterações aqui)
                    transacao_principal = Transaction.objects.create(
                        unidade_negocio=despesa.unidade_negocio,
                        description=f"Pagamento: {despesa.descricao}",
                        amount=-abs(despesa.valor),
                        category=despesa.categoria,
                        transaction_date=data_pagamento,
                        professor=despesa.professor,
                        created_by=request.user,
                    )

                    # Atualiza a despesa (sem alterações aqui)
                    despesa.status = "pago"
                    despesa.data_pagamento = data_pagamento
                    despesa.transacao = transacao_principal
                    despesa.save()

                    # Cria transação para juros/multa se houver
                    if juros_multa > 0:
                        # ==============================================================
                        # ALTERADO AQUI: Usamos get_or_create para criar a categoria se ela não existir
                        # ==============================================================
                        categoria_juros, created = Category.objects.get_or_create(
                            name="Juros e Multas Pagas",
                            type="expense",
                            unidade_negocio=despesa.unidade_negocio,
                            defaults={"tipo_dre": "despesa"},
                        )
                        # ==============================================================

                        Transaction.objects.create(
                            unidade_negocio=despesa.unidade_negocio,
                            description=f"Juros/Multa Ref: {despesa.descricao}",
                            amount=-abs(juros_multa),
                            category=categoria_juros,
                            transaction_date=data_pagamento,
                            created_by=request.user,
                        )
                        messages.success(
                            request,
                            f"Despesa e juros de R$ {juros_multa} baixados com sucesso!",
                        )
                    else:
                        messages.success(request, "Despesa baixada com sucesso!")

            except Exception as e:
                # O antigo "Category.DoesNotExist" não é mais necessário, pois o get_or_create resolve isso.
                messages.error(request, f"Ocorreu um erro ao baixar a despesa: {e}")

    return redirect("finances:despesa_list")


@admin_required
def delete_despesa_view(request, pk):
    if request.method == "POST":
        despesa = get_object_or_404(Despesa, pk=pk)
        if despesa.transacao:
            despesa.transacao.delete()
        despesa.delete()
        messages.success(request, "Despesa deletada com sucesso!")
    return redirect("finances:despesa_list")


@admin_required
def edit_despesa_view(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    if request.method == "POST":
        form = DespesaForm(request.POST, instance=despesa)
        if form.is_valid():
            try:
                with transaction.atomic():
                    despesa_atualizada = form.save()

                    if despesa_atualizada.transacao:
                        transacao = despesa_atualizada.transacao
                        transacao.description = (
                            f"Pagamento: {despesa_atualizada.descricao}"
                        )
                        transacao.amount = despesa_atualizada.valor
                        transacao.category = despesa_atualizada.categoria
                        transacao.transaction_date = (
                            despesa_atualizada.data_pagamento
                            or transacao.transaction_date
                        )
                        transacao.save()

                return JsonResponse(
                    {"status": "success", "despesa": model_to_dict(despesa_atualizada)}
                )
            except Exception as e:
                return JsonResponse({"status": "error", "errors": str(e)})
        else:
            return JsonResponse({"status": "error", "errors": form.errors})

    # GET: retorna dados para preencher o modal
    data = model_to_dict(despesa)
    return JsonResponse(data)


@admin_required
def receita_list_view(request):
    """
    View principal, agora filtrando por unidade de negócio e usando o decorador correto.
    """
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.warning(request, "Selecione uma Unidade de Negócio.")
        return redirect("scheduler:dashboard")
    
    descricao = request.GET.get('descricao', '')
    data_inicial = request.GET.get('data_inicial', '')
    data_final = request.GET.get('data_final', '')
    aluno_id = request.GET.get('aluno', '')
    categoria_id = request.GET.get('categoria', '')
    status = request.GET.get('status', '')

    receitas_list = Receita.objects.filter(unidade_negocio_id=unidade_ativa_id)

    if descricao:
        receitas_list = receitas_list.filter(descricao__icontains=descricao)
    if data_inicial:
        receitas_list = receitas_list.filter(data_competencia__gte=data_inicial)
    if data_final:
        receitas_list = receitas_list.filter(data_competencia__lte=data_final)
    if aluno_id:
        receitas_list = receitas_list.filter(aluno_id=aluno_id)
    if categoria_id:
        receitas_list = receitas_list.filter(categoria_id=categoria_id)
    if status:
        receitas_list = receitas_list.filter(status=status)

    orderby = request.GET.get('orderby', '-data_competencia') # Padrão: mais recentes
    
    # Lista de campos permitidos para evitar ordenação maliciosa
    allowed_orderby_fields = ['descricao', '-descricao', 'valor', '-valor', 'data_competencia', '-data_competencia']
    if orderby not in allowed_orderby_fields:
        orderby = '-data_competencia' # Garante um padrão seguro
        
    receitas_list = receitas_list.order_by(orderby)
    titulo = "Entradas"

    filtro_ativo = request.GET.get("filtro")
    if filtro_ativo == "a_vencer":
        hoje = now().date()
        data_limite = hoje + timedelta(days=5)
        receitas_list = receitas_list.filter(
            status="a_receber",
            data_competencia__gte=hoje,
            data_competencia__lte=data_limite,
        ).order_by("data_competencia")
        titulo = "Receitas a Vencer"

    base_qs = Receita.objects.filter(unidade_negocio_id=unidade_ativa_id)
    
    total_a_receber = base_qs.filter(status='a_receber').aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    recebido_qs = base_qs.filter(status='recebido')
    if data_inicial: recebido_qs = recebido_qs.filter(data_recebimento__gte=data_inicial)
    if data_final: recebido_qs = recebido_qs.filter(data_recebimento__lte=data_final)
    total_recebido = recebido_qs.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    
    total_recorrentes = ReceitaRecorrente.objects.filter(unidade_negocio_id=unidade_ativa_id, ativa=True).count()
    
    hoje = now().date()
    data_limite = hoje + timedelta(days=5)
    total_a_vencer = base_qs.filter(
        status='a_receber',
        data_competencia__gte=hoje,
        data_competencia__lte=data_limite
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')

    paginator = Paginator(receitas_list, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    mensalidade_form = MensalidadeReceitaForm()
    venda_form = VendaReceitaForm()

    context = {
        "page_obj": page_obj,
        "mensalidade_form": mensalidade_form,
        "venda_form": venda_form,
        "titulo": titulo,
        "orderby": orderby,
        "filtro_ativo": filtro_ativo,
        "alunos": Aluno.objects.all(),
        "categorias": Category.objects.filter(type='income'),
        "status_choices": Receita.STATUS_CHOICES,
        "filtros_aplicados": {
            'descricao': descricao,
            'data_inicial': data_inicial,
            'data_final': data_final,
            'aluno': aluno_id,
            'categoria': categoria_id,
            'status': status,
        },
        "total_a_receber": total_a_receber,
        "total_recebido": total_recebido,
        "total_recorrentes": total_recorrentes,
        "total_a_vencer": total_a_vencer,
    }
    return render(request, "finances/receita_list.html", context)


@admin_required
def baixar_receita_view(request, pk):
    receita = get_object_or_404(Receita, pk=pk)
    if request.method == "POST":
        data_recebimento_str = request.POST.get("data_recebimento")
        if data_recebimento_str:
            data_recebimento = date.fromisoformat(data_recebimento_str)

            try:
                with transaction.atomic():
                    transacao = Transaction.objects.create(
                        unidade_negocio_id=receita.unidade_negocio.id,
                        description=f"Recebimento: {receita.descricao}",
                        amount=receita.valor,
                        category=receita.categoria,
                        transaction_date=data_recebimento,
                        student=receita.aluno,
                        created_by=request.user,
                    )

                    receita.status = "recebido"
                    receita.data_recebimento = data_recebimento
                    receita.transacao = transacao
                    receita.save()

                    messages.success(request, "Receita baixada com sucesso!")

            except Exception as e:
                messages.error(request, f"Ocorreu um erro ao baixar a receita: {e}")

    return redirect("finances:receita_list")


@admin_required
def delete_receita_view(request, pk):
    if request.method == "POST":
        receita = get_object_or_404(Receita, pk=pk)
        if receita.transacao:
            receita.transacao.delete()
        receita.delete()
        messages.success(request, "Receita deletada com sucesso!")
    return redirect("finances:receita_list")


@admin_required
def edit_mensalidade(request, pk):
    """
    Busca dados e salva alterações para uma MENSALIDADE.
    AGORA USANDO A LÓGICA AJAX.
    """
    receita = get_object_or_404(
        Receita, pk=pk, produto__isnull=True
    )

    if request.method == "POST":
        form = MensalidadeReceitaForm(request.POST, instance=receita)
        if form.is_valid():
            form.save()
            # Responde com JSON em caso de sucesso
            return JsonResponse({"status": "success"})
        else:
            # Responde com JSON e os erros do formulário em caso de falha
            return JsonResponse({"status": "error", "errors": form.errors}, status=400)

    # Para GET, continua retornando os dados da receita em JSON para o modal preencher
    data = {
        "aluno": receita.aluno.pk if receita.aluno else None,
        "descricao": receita.descricao,
        "valor": str(receita.valor),
        "categoria": receita.categoria.pk if receita.categoria else None,
        "data_competencia": (
            receita.data_competencia.strftime("%Y-%m-%d")
            if receita.data_competencia
            else ""
        ),
        "data_recebimento": (
            receita.data_recebimento.strftime("%Y-%m-%d")
            if receita.data_recebimento
            else ""
        ),
    }
    return JsonResponse(data)


@admin_required
def edit_venda(request, pk):
    """
    Busca dados e salva alterações para uma VENDA.
    AGORA USANDO A LÓGICA AJAX.
    """
    receita = get_object_or_404(
        Receita, pk=pk, produto__isnull=False
    )

    quantidade_antiga = receita.quantidade
    produto_antigo = receita.produto

    if request.method == "POST":
        form = VendaReceitaForm(request.POST, instance=receita)
        if form.is_valid():
            try:
                with transaction.atomic():
                    venda_atualizada = form.save(commit=False)
                    produto_novo = venda_atualizada.produto
                    quantidade_nova = venda_atualizada.quantidade

                    if produto_antigo == produto_novo:
                        diferenca_estoque = quantidade_antiga - quantidade_nova
                        produto_novo.quantidade_em_estoque += diferenca_estoque
                    else:
                        produto_antigo.quantidade_em_estoque += quantidade_antiga
                        produto_antigo.save()
                        produto_novo.quantidade_em_estoque -= quantidade_nova

                    if produto_novo.quantidade_em_estoque < 0:
                        raise Exception(
                            f"Estoque insuficiente para o produto '{produto_novo.nome}'."
                        )

                    produto_novo.save()
                    venda_atualizada.save()
                    # Responde com JSON em caso de sucesso
                    return JsonResponse({"status": "success"})

            except Exception as e:
                # Responde com JSON em caso de erro na lógica de estoque
                return JsonResponse({"status": "error", "message": str(e)}, status=400)
        else:
            # Responde com JSON e os erros do formulário em caso de falha
            return JsonResponse({"status": "error", "errors": form.errors}, status=400)

    # Para GET, continua retornando os dados em JSON
    data = {
        "produto": receita.produto.pk if receita.produto else None,
        "quantidade": receita.quantidade,
        "descricao": receita.descricao,
        "valor": str(receita.valor),
        "categoria": receita.categoria.pk if receita.categoria else None,
        "aluno": receita.aluno.pk if receita.aluno else None,
        "data_competencia": (
            receita.data_competencia.strftime("%Y-%m-%d")
            if receita.data_competencia
            else ""
        ),
        "data_recebimento": (
            receita.data_recebimento.strftime("%Y-%m-%d")
            if receita.data_recebimento
            else ""
        ),
    }
    return JsonResponse(data)


@admin_required
def recorrencia_list_view(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if "submit_despesa" in request.POST:
        despesa_form = DespesaRecorrenteForm(request.POST)
        if despesa_form.is_valid():
            recorrente = despesa_form.save(commit=False)
            recorrente.unidade_negocio_id = unidade_ativa_id
            recorrente.save()
            messages.success(request, "Despesa recorrente salva com sucesso!")
            return redirect("finances:recorrencia_list")
    else:
        despesa_form = DespesaRecorrenteForm()

    if "submit_receita" in request.POST:
        receita_form = ReceitaRecorrenteForm(request.POST)
        if receita_form.is_valid():
            recorrente = receita_form.save(commit=False)
            recorrente.unidade_negocio_id = unidade_ativa_id
            recorrente.save()
            messages.success(request, "Receita recorrente salva com sucesso!")
            return redirect("finances:recorrencia_list")
    else:
        receita_form = ReceitaRecorrenteForm()

    despesas_recorrentes = DespesaRecorrente.objects.filter(
        unidade_negocio_id=unidade_ativa_id
    )
    receitas_recorrentes = ReceitaRecorrente.objects.filter(
        unidade_negocio_id=unidade_ativa_id
    )

    total_despesas_recorrentes = despesas_recorrentes.filter(ativa=True).aggregate(
        total=Sum('valor')
    )['total'] or Decimal('0.00')

    receitas_fixas = receitas_recorrentes.filter(ativa=True, aluno__isnull=True).aggregate(
        total=Sum('valor')
    )['total'] or Decimal('0.00')

    mensalidades = Aluno.objects.filter(
        status='ativo',
        receitas_recorrentes__unidade_negocio_id=unidade_ativa_id,
        receitas_recorrentes__ativa=True
    ).aggregate(total=Sum('valor_mensalidade'))['total'] or Decimal('0.00')

    total_receitas_recorrentes = receitas_fixas + mensalidades
    saldo_recorrente_mensal = total_receitas_recorrentes - total_despesas_recorrentes


    context = {
        "despesa_form": despesa_form,
        "receita_form": receita_form,
        "despesas_recorrentes": despesas_recorrentes,
        "receitas_recorrentes": receitas_recorrentes,
        "total_receitas_recorrentes": total_receitas_recorrentes,
        "total_despesas_recorrentes": total_despesas_recorrentes,
        "saldo_recorrente_mensal": saldo_recorrente_mensal,
    }
    return render(request, "finances/recorrencia_list.html", context)


@admin_required
def delete_despesa_recorrente_view(request, pk):
    if request.method == "POST":
        recorrente = get_object_or_404(DespesaRecorrente, pk=pk)
        recorrente.delete()
        messages.success(request, "Despesa recorrente deletada com sucesso!")
    return redirect("finances:recorrencia_list")


@admin_required
def edit_despesa_recorrente_view(request, pk):
    recorrente = get_object_or_404(DespesaRecorrente, pk=pk)
    if request.method == "POST":
        form = DespesaRecorrenteForm(request.POST, instance=recorrente)
        if form.is_valid():
            form.save()
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "errors": form.errors})
    data = model_to_dict(recorrente)
    return JsonResponse(data)


@admin_required
def delete_receita_recorrente_view(request, pk):
    if request.method == "POST":
        recorrente = get_object_or_404(ReceitaRecorrente, pk=pk)
        recorrente.delete()
        messages.success(request, "Receita recorrente deletada com sucesso!")
    return redirect("finances:recorrencia_list")


@admin_required
def edit_receita_recorrente_view(request, pk):
    recorrente = get_object_or_404(ReceitaRecorrente, pk=pk)
    if request.method == "POST":
        form = ReceitaRecorrenteForm(request.POST, instance=recorrente)
        if form.is_valid():
            form.save()
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "errors": form.errors})
    data = model_to_dict(recorrente)
    return JsonResponse(data)


def get_dre_data(unidade_negocio_id, start_date, end_date):
    """Busca e calcula os dados para o DRE de um período específico."""
    if not all([unidade_negocio_id, start_date, end_date]):
        return None

    receitas_periodo = Receita.objects.filter(
        unidade_negocio_id=unidade_negocio_id,
        data_competencia__range=[start_date, end_date],
    )
    despesas_periodo = Despesa.objects.filter(
        unidade_negocio_id=unidade_negocio_id,
        data_competencia__range=[start_date, end_date],
    )

    data = {}
    data["receitas_por_categoria"] = list(
        receitas_periodo.values("categoria__name")
        .annotate(total_cat=Sum("valor"))
        .order_by("-total_cat")
    )
    data["total_receitas"] = receitas_periodo.aggregate(total=Sum("valor"))[
        "total"
    ] or Decimal("0.00")

    custos_q = despesas_periodo.filter(categoria__tipo_dre="custo")
    despesas_q = despesas_periodo.filter(categoria__tipo_dre="despesa")

    data["custos_por_categoria"] = list(
        custos_q.values("categoria__name").annotate(total_cat=Sum("valor"))
    )
    data["total_custos"] = custos_q.aggregate(total=Sum("valor"))["total"] or Decimal(
        "0.00"
    )

    data["despesas_por_categoria"] = list(
        despesas_q.values("categoria__name").annotate(total_cat=Sum("valor"))
    )
    data["total_despesas"] = despesas_q.aggregate(total=Sum("valor"))[
        "total"
    ] or Decimal("0.00")

    data["lucro_bruto"] = data["total_receitas"] - data["total_custos"]
    data["resultado"] = data["lucro_bruto"] - data["total_despesas"]

    if data["total_receitas"] > 0:
        data["perc_custos"] = data["total_custos"] / data["total_receitas"] * 100
        data["perc_lucro_bruto"] = data["lucro_bruto"] / data["total_receitas"] * 100
        data["perc_despesas"] = data["total_despesas"] / data["total_receitas"] * 100
        data["perc_resultado"] = data["resultado"] / data["total_receitas"] * 100
    else:
        data["perc_custos"] = data["perc_lucro_bruto"] = data["perc_despesas"] = data[
            "perc_resultado"
        ] = Decimal("0.00")

    return data


def merge_and_compare_categories(principal_list, comp_list):
    """
    Combina e compara duas listas de categorias de um DRE, calculando a variação.
    Cada item da lista deve ser um dicionário com 'categoria__name' e 'total_cat'.
    """
    comp_dict = {item["categoria__name"]: item["total_cat"] for item in comp_list}
    principal_dict = {
        item["categoria__name"]: item["total_cat"] for item in principal_list
    }

    # Garante que todas as categorias de ambos os períodos sejam consideradas
    all_category_names = sorted(
        list(set(principal_dict.keys()) | set(comp_dict.keys()))
    )

    merged_list = []
    for name in all_category_names:
        principal_val = principal_dict.get(name, Decimal("0.00"))
        comp_val = comp_dict.get(name, Decimal("0.00"))

        var_abs = principal_val - comp_val
        if comp_val != 0:
            var_perc = (var_abs / comp_val) * 100
        # Se o valor comparativo for zero, qualquer valor principal é um aumento de "100%" conceitual
        elif principal_val != 0:
            var_perc = Decimal("100.0")
        else:
            var_perc = Decimal("0.0")

        merged_list.append(
            {
                "name": name,
                "principal": principal_val,
                "comparativo": comp_val,
                "var_abs": var_abs,
                "var_perc": var_perc,
            }
        )
    return merged_list


@admin_required
def dre_view(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.warning(request, "Selecione uma Unidade de Negócio.")
        return redirect("scheduler:dashboard")

    today = now().date()

    # Período Principal
    start_date = (
        date.fromisoformat(request.GET.get("start_date"))
        if request.GET.get("start_date")
        else today.replace(day=1)
    )
    end_date = (
        date.fromisoformat(request.GET.get("end_date"))
        if request.GET.get("end_date")
        else (today.replace(day=28) + timedelta(days=4)).replace(day=1)
        - timedelta(days=1)
    )

    # Período Comparativo
    start_date_comp_str = request.GET.get("start_date_comp")
    end_date_comp_str = request.GET.get("end_date_comp")
    start_date_comp = (
        date.fromisoformat(start_date_comp_str) if start_date_comp_str else None
    )
    end_date_comp = date.fromisoformat(end_date_comp_str) if end_date_comp_str else None

    dre_principal = get_dre_data(unidade_ativa_id, start_date, end_date)
    dre_comp = (
        get_dre_data(unidade_ativa_id, start_date_comp, end_date_comp)
        if start_date_comp and end_date_comp
        else None
    )

    context = {
        "start_date": start_date,
        "end_date": end_date,
        "start_date_comp": start_date_comp,
        "end_date_comp": end_date_comp,
        "dre_principal": dre_principal,
        "dre_comp": dre_comp,
    }

    # Se houver comparação, processa os dados para o template
    if dre_comp:
        variacoes = {}
        for key in [
            "total_receitas", "total_custos", "lucro_bruto",
            "total_despesas", "resultado"
        ]:
            val_principal = dre_principal.get(key, Decimal("0.00"))
            val_comp = dre_comp.get(key, Decimal("0.00"))

            var_abs = val_principal - val_comp
            var_perc = (var_abs / val_comp * 100) if val_comp != 0 else (Decimal("100.0") if val_principal != 0 else Decimal("0.0"))
            variacoes[key] = {"abs": var_abs, "perc": var_perc}
        context["variacoes"] = variacoes

        # Gera as listas mescladas para as categorias
        context["merged_receitas"] = merge_and_compare_categories(
            dre_principal.get('receitas_por_categoria', []),
            dre_comp.get('receitas_por_categoria', [])
        )
        context["merged_custos"] = merge_and_compare_categories(
            dre_principal.get('custos_por_categoria', []),
            dre_comp.get('custos_por_categoria', [])
        )
        context["merged_despesas"] = merge_and_compare_categories(
            dre_principal.get('despesas_por_categoria', []),
            dre_comp.get('despesas_por_categoria', [])
        )
    no_data = (
            dre_principal and
            dre_principal.get('total_receitas', 0) == 0 and
            dre_principal.get('total_custos', 0) == 0 and
            dre_principal.get('total_despesas', 0) == 0
        )
    context['no_data'] = no_data

    return render(request, "finances/dre_report.html", context)


@admin_required
def dre_details_view(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    category_name = request.GET.get("categoria")
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    modelo = request.GET.get("modelo")

    if not all([unidade_ativa_id, category_name, start_date_str, end_date_str, modelo]):
        messages.error(request, "Parâmetros insuficientes para gerar o detalhamento.")
        return redirect("finances:dre_report")

    try:
        start_date_obj = date.fromisoformat(start_date_str)
        end_date_obj = date.fromisoformat(end_date_str)
    except (ValueError, TypeError):
        messages.error(request, "Formato de data inválido.")
        return redirect("finances:dre_report")

    ModelClass = Receita if modelo == "receita" else Despesa
    base_qs = ModelClass.objects.filter(
        unidade_negocio_id=unidade_ativa_id,
        categoria__name=category_name,
    )
    
    if modelo == "despesa":
        classificacao = request.GET.get("classificacao")
        if classificacao:
            base_qs = base_qs.filter(categoria__tipo_dre=classificacao)

    # ===================== INÍCIO DA CORREÇÃO =====================
    # Define os campos para otimização de acordo com o modelo
    if modelo == 'receita':
        related_fields = ['aluno', 'categoria']
    else:  # 'despesa'
        related_fields = ['professor', 'categoria']
    
    lancamentos = base_qs.filter(
        data_competencia__month__gte=start_date_obj.month,
        data_competencia__day__gte=start_date_obj.day,
        data_competencia__month__lte=end_date_obj.month,
        data_competencia__day__lte=end_date_obj.day,
    ).select_related(*related_fields).order_by('-data_competencia')
    # ===================== FIM DA CORREÇÃO =====================

    lancamentos_por_ano = defaultdict(lambda: {'itens': [], 'total_ano': Decimal('0.00')})
    total_geral = Decimal('0.00')

    for lancamento in lancamentos:
        ano = lancamento.data_competencia.year
        lancamentos_por_ano[ano]['itens'].append(lancamento)
        lancamentos_por_ano[ano]['total_ano'] += lancamento.valor
        total_geral += lancamento.valor

    dados_agrupados = sorted(lancamentos_por_ano.items(), key=lambda x: x[0], reverse=True)

    context = {
        "dados_agrupados": dados_agrupados,
        "total_geral": total_geral,
        "category_name": category_name,
        "start_date": start_date_obj,
        "end_date": end_date_obj,
        "modelo": modelo,
    }
    return render(request, "finances/dre_details.html", context)


@admin_required
def export_dre_xlsx(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    start_date_comp_str = request.GET.get("start_date_comp")
    end_date_comp_str = request.GET.get("end_date_comp")

    start_date = date.fromisoformat(start_date_str)
    end_date = date.fromisoformat(end_date_str)
    dre_data = get_dre_data(unidade_ativa_id, start_date, end_date)
    
    dre_comp = None
    if start_date_comp_str and end_date_comp_str:
        start_date_comp = date.fromisoformat(start_date_comp_str)
        end_date_comp = date.fromisoformat(end_date_comp_str)
        dre_comp = get_dre_data(unidade_ativa_id, start_date_comp, end_date_comp)

    wb = Workbook()
    ws = wb.active
    ws.title = "DRE"

    # --- Estilos (sem alterações) ---
    bold_font = Font(bold=True)
    white_font_bold = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    total_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    final_fill = PatternFill(start_color="424242", end_color="424242", fill_type="solid")
    right_align = Alignment(horizontal="right")
    center_align = Alignment(horizontal="center")

    headers = ['Descrição', 'Período Principal']
    if dre_comp:
        headers.extend(['Período Comparativo', 'Variação (R$)', 'Variação (%)'])
    num_cols = len(headers)

    # --- Cabeçalho do Arquivo (sem alterações) ---
    ws['A1'] = 'Demonstrativo de Resultados (DRE)'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    periodo_str = f'Principal: {start_date.strftime("%d/%m/%Y")} a {end_date.strftime("%d/%m/%Y")}'
    if dre_comp:
        periodo_str += f' | Comparativo: {start_date_comp.strftime("%d/%m/%Y")} a {end_date_comp.strftime("%d/%m/%Y")}'
    ws['A2'] = periodo_str
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'].alignment = center_align

    # --- Cabeçalho da Tabela (sem alterações) ---
    for col_idx, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header_title
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = right_align if col_idx > 1 else Alignment(horizontal="left")

    row_cursor = 5

    def apply_row_styles(row_idx, is_total=False, is_final=False):
        fill = total_fill if is_total else (final_fill if is_final else None)
        font = bold_font if is_total else (white_font_bold if is_final else None)
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill: cell.fill = fill
            if font: cell.font = font
            if col_idx > 1: cell.alignment = right_align

    # ===================== INÍCIO DA CORREÇÃO =====================
    def add_data_row(data, indent=0):
        nonlocal row_cursor
        ws.cell(row=row_cursor, column=1).value = f'{"  " * indent}{data[0]}'
        for i, value in enumerate(data[1:], 2):
            cell = ws.cell(row=row_cursor, column=i)
            cell.value = value
            if isinstance(value, (Decimal, float)):
                # Lógica corrigida: só formata como % se for a última coluna E houver comparação
                if dre_comp and i == num_cols:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = 'R$ #,##0.00'
        row_cursor += 1
    # ===================== FIM DA CORREÇÃO =====================
    
    # --- Lógica para preencher os dados (sem alterações) ---
    if dre_comp:
        variacoes = {}
        for key in ['total_receitas', 'total_custos', 'lucro_bruto', 'total_despesas', 'resultado']:
            val_principal = dre_data.get(key, Decimal('0.00'))
            val_comp = dre_comp.get(key, Decimal('0.00'))
            var_abs = val_principal - val_comp
            var_perc = (var_abs / val_comp * 100) if val_comp != 0 else (Decimal("100.0") if val_principal != 0 else Decimal("0.0"))
            variacoes[key] = {'abs': var_abs, 'perc': var_perc}
        
        merged_receitas = merge_and_compare_categories(dre_data.get('receitas_por_categoria', []), dre_comp.get('receitas_por_categoria', []))
        merged_custos = merge_and_compare_categories(dre_data.get('custos_por_categoria', []), dre_comp.get('custos_por_categoria', []))
        merged_despesas = merge_and_compare_categories(dre_data.get('despesas_por_categoria', []), dre_comp.get('despesas_por_categoria', []))
        
        add_data_row(['(+) Receita Operacional Bruta', dre_data['total_receitas'], dre_comp['total_receitas'], variacoes['total_receitas']['abs'], variacoes['total_receitas']['perc']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for r in merged_receitas: add_data_row([f"+ {r['name']}", r['principal'], r['comparativo'], r['var_abs'], r['var_perc']], indent=1)
        
        add_data_row(['(-) Custos Diretos', -dre_data['total_custos'], -dre_comp['total_custos'], -variacoes['total_custos']['abs'], -variacoes['total_custos']['perc']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for c in merged_custos: add_data_row([f"- {c['name']}", -c['principal'], -c['comparativo'], -c['var_abs'], -c['var_perc']], indent=1)
        
        add_data_row(['(=) Lucro Bruto', dre_data['lucro_bruto'], dre_comp['lucro_bruto'], variacoes['lucro_bruto']['abs'], variacoes['lucro_bruto']['perc']])
        apply_row_styles(row_cursor - 1, is_total=True)

        add_data_row(['(-) Despesas Operacionais', -dre_data['total_despesas'], -dre_comp['total_despesas'], -variacoes['total_despesas']['abs'], -variacoes['total_despesas']['perc']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for d in merged_despesas: add_data_row([f"- {d['name']}", -d['principal'], -d['comparativo'], -d['var_abs'], -d['var_perc']], indent=1)

        add_data_row(['(=) Resultado do Período', dre_data['resultado'], dre_comp['resultado'], variacoes['resultado']['abs'], variacoes['resultado']['perc']])
        apply_row_styles(row_cursor - 1, is_final=True)

    else: # Lógica para período único
        add_data_row(['(+) Receita Operacional Bruta', dre_data['total_receitas']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for r in dre_data['receitas_por_categoria']: add_data_row([f"+ {r['categoria__name']}", r['total_cat']], indent=1)
        
        add_data_row(['(-) Custos Diretos', -dre_data['total_custos']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for c in dre_data['custos_por_categoria']: add_data_row([f"- {c['categoria__name']}", -c['total_cat']], indent=1)

        add_data_row(['(=) Lucro Bruto', dre_data['lucro_bruto']])
        apply_row_styles(row_cursor - 1, is_total=True)

        add_data_row(['(-) Despesas Operacionais', -dre_data['total_despesas']])
        apply_row_styles(row_cursor - 1, is_total=True)
        for d in dre_data['despesas_por_categoria']: add_data_row([f"- {d['categoria__name']}", -d['total_cat']], indent=1)

        add_data_row(['(=) Resultado do Período', dre_data['resultado']])
        apply_row_styles(row_cursor - 1, is_final=True)

    # --- Ajuste final de colunas e cores (sem alterações) ---
    ws.column_dimensions['A'].width = 50
    for i in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    if dre_comp:
        file_name = (
            f"DRE Comparativo "
            f"{start_date.strftime('%d-%m-%Y')} a {end_date.strftime('%d-%m-%Y')}"
            f" vs {start_date_comp.strftime('%d-%m-%Y')} a {end_date_comp.strftime('%d-%m-%Y')}.xlsx"
        )
    else:
        file_name = (
            f"DRE "
            f"{start_date.strftime('%d-%m-%Y')} a {end_date.strftime('%d-%m-%Y')}.xlsx"
        )
    
    # Resposta HTTP com o nome de arquivo correto e sem linhas duplicadas
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    wb.save(response)
    return response


@admin_required
def export_dre_pdf(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    start_date_comp_str = request.GET.get("start_date_comp")
    end_date_comp_str = request.GET.get("end_date_comp")

    start_date = date.fromisoformat(start_date_str)
    end_date = date.fromisoformat(end_date_str)
    # Renomeando 'dre_data' para 'dre_principal' para alinhar com o template
    dre_principal = get_dre_data(unidade_ativa_id, start_date, end_date)
    
    dre_comp = None
    start_date_comp = None
    end_date_comp = None
    if start_date_comp_str and end_date_comp_str:
        start_date_comp = date.fromisoformat(start_date_comp_str)
        end_date_comp = date.fromisoformat(end_date_comp_str)
        dre_comp = get_dre_data(unidade_ativa_id, start_date_comp, end_date_comp)

    # Contexto agora usa 'dre_principal', que é o nome esperado pelo template
    context = {
        "start_date": start_date,
        "end_date": end_date,
        "dre_principal": dre_principal,
        "dre_comp": dre_comp,
        "start_date_comp": start_date_comp,
        "end_date_comp": end_date_comp,
    }

    if dre_comp:
        variacoes = {}
        for key in ["total_receitas", "total_custos", "lucro_bruto", "total_despesas", "resultado"]:
            val_principal = dre_principal.get(key, Decimal("0.00"))
            val_comp = dre_comp.get(key, Decimal("0.00"))
            var_abs = val_principal - val_comp
            var_perc = (var_abs / val_comp * 100) if val_comp != 0 else (Decimal("100.0") if val_principal != 0 else Decimal("0.0"))
            variacoes[key] = {"abs": var_abs, "perc": var_perc}
        context["variacoes"] = variacoes

        # Adicionando as listas detalhadas ao contexto (a parte que estava faltando)
        context["merged_receitas"] = merge_and_compare_categories(
            dre_principal.get('receitas_por_categoria', []), dre_comp.get('receitas_por_categoria', [])
        )
        context["merged_custos"] = merge_and_compare_categories(
            dre_principal.get('custos_por_categoria', []), dre_comp.get('custos_por_categoria', [])
        )
        context["merged_despesas"] = merge_and_compare_categories(
            dre_principal.get('despesas_por_categoria', []), dre_comp.get('despesas_por_categoria', [])
        )

    html_string = render_to_string("finances/dre_pdf_template.html", context)
    response = HttpResponse(content_type="application/pdf")
    
    def format_date_for_filename(d):
        # Converte date para dd-mm-yyyy
        return d.strftime("%d-%m-%Y")

    if dre_comp:
        file_name = (
            f"DRE Comparativo {format_date_for_filename(start_date)} a {format_date_for_filename(end_date)} "
            f"vs {format_date_for_filename(start_date_comp)} a {format_date_for_filename(end_date_comp)}.pdf"
        )
    else:
        file_name = f"DRE {format_date_for_filename(start_date)} a {format_date_for_filename(end_date)}.pdf"

    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse("Ocorreu um erro ao gerar o PDF.")
    return response


@login_required
def get_aluno_details(request, aluno_id):
    """
    Retorna JSON com detalhes do aluno para preenchimento automático
    no formulário de Contas a Receber.
    """
    try:
        aluno = get_object_or_404(Aluno, pk=aluno_id)
        status_pagamento = aluno.get_status_pagamento()
        data = {
            "id": aluno.id,
            "nome_completo": aluno.nome_completo,
            "valor_mensalidade": float(aluno.valor_mensalidade or 0),
            "dia_vencimento": aluno.dia_vencimento or "",
            "status": status_pagamento["status"],
            "status_cor": status_pagamento["cor"],
            "email": aluno.email or "",
            "telefone": aluno.telefone or "",
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@admin_required
@require_POST
def toggle_recorrencia_ativa(request):
    try:
        model_name = request.POST.get("model")
        pk = request.POST.get("pk")
        is_active = request.POST.get("is_active") == "true"

        Model = apps.get_model("finances", model_name)
        recorrencia = get_object_or_404(Model, pk=pk)

        recorrencia.ativa = is_active
        recorrencia.save()

        return JsonResponse(
            {"status": "success", "message": f"{model_name} atualizada."}
        )
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)


@admin_required
def calcular_pagamento_professor_ajax(request):
    professor_id = request.GET.get("professor_id")
    data_inicial_str = request.GET.get("data_inicial")
    data_final_str = request.GET.get("data_final")

    if not all([professor_id, data_inicial_str, data_final_str]):
        return JsonResponse(
            {"status": "error", "message": "Parâmetros ausentes."}, status=400
        )

    try:
        professor = get_object_or_404(CustomUser, pk=professor_id)
        data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
        data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
    except (ValueError, CustomUser.DoesNotExist):
        return JsonResponse(
            {"status": "error", "message": "Parâmetros inválidos."}, status=400
        )

    # 1. Filtra aulas realizadas
    q_realizadas_normal = Q(
        status="Realizada", relatorioaula__professor_que_validou=professor
    ) & ~Q(modalidade__nome__icontains="atividade complementar")
    q_realizadas_ac = Q(
        status="Realizada",
        modalidade__nome__icontains="atividade complementar",
        presencas_professores__professor=professor,
        presencas_professores__status="presente",
    )

    aulas_realizadas = (
        Aula.objects.filter(
            data_hora__date__gte=data_inicial, data_hora__date__lte=data_final
        )
        .filter(q_realizadas_normal | q_realizadas_ac)
        .distinct()
    )

    # 2. Agrupa por modalidade
    calculo_detalhado = []
    modalidades_envolvidas = Modalidade.objects.filter(
        aula__in=aulas_realizadas
    ).distinct()

    for modalidade in modalidades_envolvidas:
        aulas_nesta_modalidade = aulas_realizadas.filter(modalidade=modalidade)

        item_calculo = {
            "modalidade_nome": modalidade.nome,
            "tipo_pagamento_display": modalidade.get_tipo_pagamento_display(),
            "valor_unitario": float(modalidade.valor_pagamento_professor),
        }

        # 3. Aplica a regra
        if modalidade.tipo_pagamento == "aluno":
            quantidade = PresencaAluno.objects.filter(
                aula__in=aulas_nesta_modalidade, status="presente"
            ).count()
            item_calculo["unidade_contagem"] = "aluno(s) presente(s)"
        else:
            quantidade = aulas_nesta_modalidade.count()
            item_calculo["unidade_contagem"] = "aula(s)"

        item_calculo["quantidade"] = quantidade
        item_calculo["subtotal"] = float(
            Decimal(quantidade) * modalidade.valor_pagamento_professor
        )

        if quantidade > 0:
            calculo_detalhado.append(item_calculo)

    return JsonResponse(
        {
            "status": "success",
            "professor_name": professor.get_full_name() or professor.username,
            "calculo": calculo_detalhado,
        }
    )


def _process_aging_data(queryset, date_field_name, today):
    """
    Função auxiliar para processar uma queryset de contas (Receita ou Despesa)
    e retornar os dados agrupados por entidade e por faixa de vencimento.
    """
    # --- CORRIGIDO: Chaves do dicionário renomeadas para serem válidas no template ---
    buckets = {
        'a_vencer': {'min': -9999, 'max': 0,    'label': 'A Vencer'},
        'd1_30':    {'min': 1,    'max': 30,   'label': '1-30 dias'},
        'd31_60':   {'min': 31,   'max': 60,   'label': '31-60 dias'},
        'd61_90':   {'min': 61,   'max': 90,   'label': '61-90 dias'},
        'd90_plus': {'min': 91,   'max': 9999, 'label': '90+ dias'},
    }
    
    entity_data = defaultdict(lambda: {
        'entity': None,
        'total_geral': Decimal('0.00'),
        'buckets': {key: Decimal('0.00') for key in buckets}
    })

    column_totals = {key: Decimal('0.00') for key in buckets}
    grand_total = Decimal('0.00')

    for item in queryset:
        days_overdue = (today - getattr(item, date_field_name)).days
        
        for key, limits in buckets.items():
            if limits['min'] <= days_overdue <= limits['max']:
                bucket_key = key
                break
        else:
            continue

        if isinstance(item, Receita) and item.aluno:
            entity_id = item.aluno.id
            entity_name = item.aluno.nome_completo
        else:
            entity_id = item.descricao.lower()
            entity_name = item.descricao
        
        entity_data[entity_id]['entity'] = entity_name
        entity_data[entity_id]['buckets'][bucket_key] += item.valor
        entity_data[entity_id]['total_geral'] += item.valor
        
        column_totals[bucket_key] += item.valor
        grand_total += item.valor

    sorted_entities = sorted(entity_data.values(), key=lambda x: x['total_geral'], reverse=True)

    return {
        'entities': sorted_entities,
        'column_totals': column_totals,
        'grand_total': grand_total,
        'bucket_labels': [b['label'] for b in buckets.values()]
    }


@admin_required
def aging_report_view(request):
    unidade_ativa_id = request.session.get("unidade_ativa_id")
    if not unidade_ativa_id:
        messages.warning(request, "Selecione uma Unidade de Negócio.")
        return redirect("scheduler:dashboard")

    today = now().date()

    # Processa Contas a Receber
    recebiveis_abertos = Receita.objects.filter(
        unidade_negocio_id=unidade_ativa_id,
        status='a_receber'
    ).select_related('aluno')
    aging_recebiveis = _process_aging_data(recebiveis_abertos, 'data_competencia', today)

    # Processa Contas a Pagar
    pagaveis_abertos = Despesa.objects.filter(
        unidade_negocio_id=unidade_ativa_id,
        status='a_pagar'
    )
    aging_pagaveis = _process_aging_data(pagaveis_abertos, 'data_competencia', today)

    context = {
        "aging_recebiveis": aging_recebiveis,
        "aging_pagaveis": aging_pagaveis,
        "report_date": today,
    }
    
    return render(request, 'finances/aging_report.html', context)

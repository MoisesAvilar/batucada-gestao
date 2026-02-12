
import os
import io
import base64
import re
import matplotlib
from django.conf import settings
import matplotlib.pyplot as plt
import google.generativeai as genai
import json
import markdown
from django.db import transaction, DatabaseError
from django.core.exceptions import ValidationError
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
from leads.models import Lead, InteracaoLead
from .models import (
    Aula,
    Aluno,
    RelatorioAula,
    Modalidade,
    CustomUser,
    PresencaAluno,
    PresencaProfessor,
    ItemRudimento,
    TourVisto,
)
from finances.models import ReceitaRecorrente, Category, Receita, Transaction
from django.utils import timezone

# --- IMPORTS ATUALIZADOS ---
from django.forms import formset_factory
from .forms import (
    AlunoForm,
    AlunoChoiceForm,
    AulaForm,
    ModalidadeForm,
    ProfessorForm,
    ProfessorChoiceForm,
    RelatorioAulaForm,  # O formulário principal, agora menor
    ItemRudimentoFormSet,  # O novo formset de rudimentos
    ItemRitmoFormSet,  # O novo formset de ritmo
    ItemViradaFormSet,  # O novo formset de viradas
    UserProfileForm,
    PresencaAlunoFormSet,
    PresencaProfessorFormSet,
)
from django.contrib import messages
import calendar
from datetime import datetime, date, timedelta
from django.db.models import Count, Min, Case, When, Q, OuterRef, Subquery, F, Value as V
from django.db.models.functions import TruncMonth, Coalesce
from django.core.paginator import Paginator
from collections import defaultdict
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from decimal import Decimal

# --- NOVOS IMPORTS PARA O EXCEL ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
matplotlib.use('Agg')


GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)


# --- Funções de Teste para Permissões ---
def is_admin(user):
    return user.is_authenticated and user.tipo == "admin"


# --- Função auxiliar para verificar conflitos (NÃO É UMA VIEW) ---
def _check_conflito_aula(professor_ids, data_hora, aula_id=None):
    if not professor_ids:
        return {"conflito": False, "mensagem": "Horário disponível."}
    aulas_conflitantes = Aula.objects.filter(
        professores__id__in=professor_ids, data_hora=data_hora
    )
    if aula_id:
        aulas_conflitantes = aulas_conflitantes.exclude(pk=aula_id)
    if aulas_conflitantes.exists():
        # ... (lógica de mensagem de erro) ...
        return {"conflito": True, "mensagem": "Conflito de horário detectado."}
    return {"conflito": False, "mensagem": "Horário disponível."}


def _check_conflito_aluno(aluno_ids, data_hora, aula_id=None):
    """
    Verifica se algum dos alunos já tem uma aula agendada no horário especificado.
    Ignora aulas canceladas ou que já foram repostas.
    """
    if not aluno_ids:
        return {"conflito": False}

    status_de_conflito = ["Agendada", "Realizada", "Aluno Ausente"]

    # ★ MUDANÇA: Adicionamos prefetch_related para performance ★
    aulas_conflitantes = Aula.objects.filter(
        alunos__id__in=aluno_ids,
        data_hora=data_hora,
        status__in=status_de_conflito
    ).prefetch_related('alunos', 'professores').distinct() # Adicionado prefetch

    if aula_id:
        aulas_conflitantes = aulas_conflitantes.exclude(pk=aula_id)

    aula_conflitante = aulas_conflitantes.first()

    if aula_conflitante:
        aluno_nome = "Um dos alunos"
        try:
            ids_na_aula = aula_conflitante.alunos.values_list('id', flat=True)
            id_comum = set(aluno_ids).intersection(ids_na_aula).pop()
            aluno = Aluno.objects.get(id=id_comum)
            aluno_nome = aluno.nome_completo
        except:
            pass 

        # ★ NOVO: Pega os nomes dos professores da aula conflitante ★
        prof_nomes = [p.username.title() for p in aula_conflitante.professores.all()]
        professores_conflito = ", ".join(prof_nomes) if prof_nomes else "N/A"

        return {
            "conflito": True,
            "aluno_nome": aluno_nome,
            "aula_conflitante_pk": aula_conflitante.pk,
            "aula_conflitante_nome": str(aula_conflitante),
            "professores_conflito": professores_conflito, # ★ NOVO CAMPO RETORNADO ★
        }
    
    return {"conflito": False}


# --- Views Principais (dashboard) ---
@login_required
@never_cache
def dashboard(request):
    now = timezone.now()
    today = now.date()

    # Aulas pendentes fixas no usuário logado (para ação imediata)
    aulas_pendentes_validacao = Aula.objects.filter(
        professores=request.user, status="Agendada", data_hora__lt=now
    ).order_by("data_hora")
    aulas_pendentes_count = aulas_pendentes_validacao.count()

    today_iso = today.strftime("%Y-%m-%d")
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    week_start_iso = start_of_week.strftime("%Y-%m-%d")
    week_end_iso = end_of_week.strftime("%Y-%m-%d")

    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    # Captura o filtro do professor (ex: Ana)
    professor_filtro_id = request.GET.get("professor_filtro_id")

    # --- Lógica Específica por Tipo de Usuário ---
    if request.user.tipo in ["admin", "comercial"]:
        # 1. Definição do QuerySet Base com Filtro (Ajuste aqui)
        aulas_qs = Aula.objects.all()
        if professor_filtro_id and professor_filtro_id.isdigit():
            aulas_qs = aulas_qs.filter(professores__id=professor_filtro_id).distinct()
        
        # 2. Stats Operacionais (KPIs fixos da unidade, não mudam com o filtro do calendário)
        aulas_hoje_count = Aula.objects.filter(data_hora__date=today).count()
        aulas_semana_count = Aula.objects.filter(
            data_hora__date__range=[start_of_week, end_of_week]
        ).count()
        aulas_agendadas_total = Aula.objects.filter(
            status="Agendada", data_hora__gte=now
        ).count()
        novos_alunos_mes = Aluno.objects.filter(
            data_criacao__year=today.year, data_criacao__month=today.month
        ).count()

        # 3. Stats Financeiros (Mantidos conforme sua lógica original)
        kpi_recebido = kpi_atrasado = kpi_em_aberto = kpi_previsto = Decimal("0.00")
        count_recebido = count_atrasado = count_em_aberto = count_previsto = 0
        kpi_mes_ref, kpi_ano_ref = today.month, today.year
        
        unidade_id = request.session.get("unidade_ativa_id")
        if unidade_id:
            cat_mensalidade = Category.objects.filter(name__iexact="Mensalidade", unidade_negocio_id=unidade_id).first()
            if cat_mensalidade:
                receitas_dict = {r.aluno_id: r for r in Receita.objects.filter(
                    unidade_negocio_id=unidade_id, categoria=cat_mensalidade,
                    data_competencia__year=kpi_ano_ref, data_competencia__month=kpi_mes_ref
                ).select_related('transacao')}
                
                transacoes_dict = {t.student_id: t for t in Transaction.objects.filter(
                    unidade_negocio_id=unidade_id, category=cat_mensalidade,
                    transaction_date__year=kpi_ano_ref, transaction_date__month=kpi_mes_ref,
                    receita__isnull=True, student__isnull=False
                )}

                for aluno in Aluno.objects.filter(status="ativo"):
                    valor, status = Decimal("0.00"), ""
                    if aluno.id in receitas_dict:
                        r = receitas_dict[aluno.id]
                        valor, status = r.valor, ('Paga' if r.status == 'recebido' or r.transacao else 'Pendente')
                        if status == 'Pendente':
                            venc = r.data_recebimento or (date(kpi_ano_ref, kpi_mes_ref, aluno.dia_vencimento) if aluno.dia_vencimento else date(kpi_ano_ref, kpi_mes_ref, 28))
                            status = 'Atrasada' if today > venc else 'Em aberto'
                    elif aluno.id in transacoes_dict:
                        valor, status = transacoes_dict[aluno.id].amount, 'Paga'
                    elif aluno.valor_mensalidade and aluno.dia_vencimento:
                        valor = aluno.valor_mensalidade
                        venc = date(kpi_ano_ref, kpi_mes_ref, aluno.dia_vencimento)
                        status = 'Atrasada' if today > venc else 'Em aberto'
                    
                    if status == 'Paga': kpi_recebido += valor; count_recebido += 1
                    elif status == 'Atrasada': kpi_atrasado += valor; count_atrasado += 1
                    elif status == 'Em aberto': kpi_em_aberto += valor; count_em_aberto += 1

                kpi_previsto = kpi_recebido + kpi_atrasado + kpi_em_aberto
                count_previsto = count_recebido + count_atrasado + count_em_aberto

        contexto = {
            "titulo": f"Painel de Controle - {request.user.get_tipo_display()}",
            "aulas_hoje_count": aulas_hoje_count, "aulas_semana_count": aulas_semana_count,
            "aulas_agendadas_total": aulas_agendadas_total, "novos_alunos_mes": novos_alunos_mes,
            "professores_list": CustomUser.objects.filter(tipo__in=["professor", "admin", "comercial"], is_active=True).order_by("first_name"),
            "primeiro_dia_mes": today.replace(day=1).strftime("%Y-%m-%d"),
            "ultimo_dia_mes": today.replace(day=calendar.monthrange(today.year, today.month)[1]).strftime("%Y-%m-%d"),
            "aulas_pendentes_validacao": aulas_pendentes_validacao,
            "kpi_recebido": kpi_recebido, "count_recebido": count_recebido,
            "kpi_atrasado": kpi_atrasado, "count_atrasado": count_atrasado,
            "kpi_em_aberto": kpi_em_aberto, "count_em_aberto": count_em_aberto,
            "kpi_previsto": kpi_previsto, "count_previsto": count_previsto,
            "kpi_mes_ref": kpi_mes_ref, "kpi_ano_ref": kpi_ano_ref,
        }
        ja_viu_tour = request.user.tours_vistos.filter(tour_id="horarios_fixos_v1").exists()
        contexto["mostrar_tour_horarios"] = not ja_viu_tour

    else:  # Professor
        aulas_qs = Aula.objects.filter(professores=request.user).distinct()
        contexto = {
            "titulo": "Painel de Controle",
            "aulas_hoje_count": aulas_qs.filter(data_hora__date=today).count(),
            "aulas_semana_count": aulas_qs.filter(data_hora__date__range=[start_of_week, end_of_week]).count(),
            "aulas_pendentes_count": aulas_pendentes_count,
            "aulas_pendentes_validacao": aulas_pendentes_validacao,
            "mostrar_tour_horarios": False,
        }

    # --- Lógica do Calendário e Agenda Lateral (Agora usa o aulas_qs filtrado acima) ---
    aulas_do_mes = (
        aulas_qs.filter(data_hora__year=year, data_hora__month=month)
        .select_related("modalidade")
        .prefetch_related("alunos", "professores")
        .order_by("data_hora")
    )
    
    aulas_por_dia = defaultdict(list)
    for aula in aulas_do_mes:
        aulas_por_dia[localtime(aula.data_hora).day].append(aula)

    cal = calendar.Calendar(firstweekday=6)
    semanas_do_mes = cal.monthdayscalendar(year, month)
    calendario_final = []
    for semana in semanas_do_mes:
        semana_com_aulas = []
        for dia in semana:
            semana_com_aulas.append({"dia": dia, "aulas": aulas_por_dia.get(dia, [])})
        calendario_final.append(semana_com_aulas)

    AlunoFormSetModal = formset_factory(AlunoChoiceForm, extra=1, can_delete=False)
    ProfessorFormSetModal = formset_factory(ProfessorChoiceForm, extra=1, can_delete=False)

    contexto.update({
        "today": today,
        "mes_atual": date(year, month, 1),
        "calendario_mes": calendario_final,
        "aulas_do_mes_lista": aulas_do_mes, # Agora filtrada pela Ana se selecionada
        "today_iso": today_iso,
        "week_start_iso": week_start_iso,
        "week_end_iso": week_end_iso,
        "aula_form_modal": AulaForm(user=request.user),
        "aluno_formset_modal": AlunoFormSetModal(prefix="alunos"),
        "professor_formset_modal": ProfessorFormSetModal(prefix="professores"),
        "form_action_modal": reverse("scheduler:aula_agendar"),
    })

    return render(request, "scheduler/dashboard.html", contexto)

@login_required
def get_calendario_html(request):
    try:
        year = int(request.GET.get("year"))
        month = int(request.GET.get("month"))
    except (ValueError, TypeError):
        return HttpResponse("Parâmetros de ano/mês inválidos.", status=400)

    # ★★★ LÓGICA DE FILTRO ATUALIZADA ★★★
    professor_filtro_id = request.GET.get("professor_filtro_id")

    if request.user.tipo in ["admin", "comercial"]:
        aulas_qs = Aula.objects.all()
        if professor_filtro_id and professor_filtro_id.isdigit():
            # Usamos distinct() para evitar duplicatas e garantimos que o ID é válido
            aulas_qs = aulas_qs.filter(professores__id=professor_filtro_id).distinct()
    else:
        aulas_qs = Aula.objects.filter(professores=request.user).distinct()

    # O resto da função permanece igual
    aulas_do_mes = (
        aulas_qs.filter(data_hora__year=year, data_hora__month=month)
        .select_related("modalidade")
        .prefetch_related("alunos", "professores")
        .order_by("data_hora")
    )

    aulas_por_dia = defaultdict(list)
    for aula in aulas_do_mes:
        aulas_por_dia[localtime(aula.data_hora).day].append(aula)

    cal = calendar.Calendar(firstweekday=6)
    semanas_do_mes = cal.monthdayscalendar(year, month)
    calendario_final = []
    for semana in semanas_do_mes:
        semana_com_aulas = []
        for dia in semana:
            semana_com_aulas.append({"dia": dia, "aulas": aulas_por_dia.get(dia, [])})
        calendario_final.append(semana_com_aulas)

    contexto = {
        "calendario_mes": calendario_final,
        "aulas_do_mes_lista": aulas_do_mes,
        "mes_atual": date(year, month, 1),
        "today": date.today(),
    }

    html = render_to_string(
        "scheduler/partials/calendario_content.html", contexto, request=request
    )
    return HttpResponse(html)


# --- Views de Aulas (agendar_aula, editar_aula, excluir_aula) ---
@login_required
def agendar_aula(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    AlunoFormSet = formset_factory(AlunoChoiceForm, extra=1, can_delete=True)
    ProfessorFormSet = formset_factory(ProfessorChoiceForm, extra=1, can_delete=True)

    presenca_original_id = request.GET.get("reposicao_de")
    presenca_original = None
    initial_form_data = {}
    initial_aluno_data = []

    if presenca_original_id:
        try:
            presenca_original = PresencaAluno.objects.select_related(
                "aula", "aluno", "aula__modalidade"
            ).get(
                id=presenca_original_id,
                status="ausente",
                tipo_falta="justificada",
                aula_reposicao__isnull=True,
            )
            initial_form_data["modalidade"] = presenca_original.aula.modalidade
            initial_aluno_data.append({"aluno": presenca_original.aluno})
        except PresencaAluno.DoesNotExist:
            messages.error(request, "Reposição não encontrada ou já agendada.")
            presenca_original_id = None

    if request.method == "POST":
        # Passamos o usuário para o formulário, como definido na solução anterior
        form = AulaForm(request.POST, user=request.user)
        aluno_formset = AlunoFormSet(request.POST, prefix="alunos")

        professor_formset_is_valid = False
        if request.user.tipo == "admin":
            professor_formset = ProfessorFormSet(request.POST, prefix="professores")
            if professor_formset.is_valid():
                professor_formset_is_valid = True
        else:
            professor_formset = ProfessorFormSet(prefix="professores")
            professor_formset_is_valid = True

        if form.is_valid() and aluno_formset.is_valid() and professor_formset_is_valid:
            modalidade = form.cleaned_data.get("modalidade")
            alunos_ids = {
                f["aluno"].id
                for f in aluno_formset.cleaned_data
                if f and not f.get("DELETE")
            }
            # O formulário agora lida com o fuso horário automaticamente.
            data_hora_inicial = form.cleaned_data.get("data_hora")

            is_recorrente = form.cleaned_data.get("recorrente_mensal")

            # --- ★★★ INÍCIO DA CORREÇÃO ★★★ ---
            # A lógica para obter o status agora considera o tipo de usuário.
            if request.user.tipo == "professor":
                # Se for professor, o status é sempre 'Agendada' na criação.
                status = "Agendada"
            else:
                # Se for admin, pegamos o status do formulário (que já foi filtrado).
                # Adicionamos 'Agendada' como um valor padrão por segurança.
                status = form.cleaned_data.get("status", "Agendada")
            # --- ★★★ FIM DA CORREÇÃO ★★★ ---

            if request.user.tipo == "admin":
                professores_ids = {
                    f["professor"].id
                    for f in professor_formset.cleaned_data
                    if f and not f.get("DELETE")
                }
            else:
                professores_ids = {request.user.id}

            is_ac = "atividade complementar" in modalidade.nome.lower()

            if (is_ac and not professores_ids) or (
                not is_ac and (not alunos_ids or not professores_ids)
            ):
                error_message = (
                    "Para Atividade Complementar, é obrigatório selecionar pelo menos um professor."
                    if is_ac
                    else "Para aulas normais, é obrigatório selecionar pelo menos um aluno E um professor."
                )
                if is_ajax:
                    return JsonResponse(
                        {"success": False, "errors": [error_message]}, status=400
                    )
                else:
                    messages.error(request, error_message)

            else:
                datas_para_agendar = []

                raw_id = request.POST.get("reposicao_de_id") or request.GET.get("reposicao_de")
                reposicao_de_id_hidden = None
                reposicao_de_id_hidden = request.POST.get("reposicao_de_id") or request.GET.get("reposicao_de")
                
                if raw_id:
                    reposicao_de_id_hidden = str(raw_id).replace('.', '')
                    is_recorrente = False

                if is_recorrente:
                    dia_semana = data_hora_inicial.weekday()
                    mes, ano = data_hora_inicial.month, data_hora_inicial.year
                    cal = calendar.Calendar()
                    for dia in cal.itermonthdates(ano, mes):
                        if (
                            dia.month == mes
                            and dia.weekday() == dia_semana
                            and dia >= data_hora_inicial.date()
                        ):
                            datas_para_agendar.append(
                                data_hora_inicial.replace(
                                    year=dia.year, month=dia.month, day=dia.day
                                )
                            )
                else:
                    datas_para_agendar.append(data_hora_inicial)

                conflitos_encontrados = []
                for data_agendamento in datas_para_agendar:
                    # 1. Checagem de professor (já existe)
                    conflito_info_prof = _check_conflito_aula(
                        list(professores_ids), data_agendamento
                    )
                    if conflito_info_prof["conflito"]:
                        mensagem = conflito_info_prof.get("mensagem", "Conflito de horário de professor")
                        conflitos_encontrados.append(
                            f"{mensagem} na data {data_agendamento.strftime('%d/%m')}."
                        )
                        continue  # Se o prof não pode, nem checa o aluno

                    if not is_ac: 
                        conflito_info_aluno = _check_conflito_aluno(
                            list(alunos_ids), data_agendamento
                        )
                        if conflito_info_aluno["conflito"]:
                            aluno_nome = conflito_info_aluno.get('aluno_nome', 'Um dos alunos')
                            aula_pk = conflito_info_aluno.get('aula_conflitante_pk')
                            prof_nomes = conflito_info_aluno.get('professores_conflito', 'N/A')
                            
                            # Constrói a mensagem
                            link_aula = reverse('scheduler:aula_validar', args=[aula_pk])
                            # ★ LINK CORRIGIDO ★
                            link_substituicao = reverse('scheduler:aulas_para_substituir') 
                            
                            mensagem_html = (
                                f"<b>Conflito:</b> O aluno <strong>{aluno_nome}</strong> já tem uma aula com <strong>{prof_nomes}</strong> neste horário. "
                                f"<a href='{link_aula}' target='_blank' class='alert-link'>Clique para substituir essa aula</a>."
                                f"<br><small>Caso queira substituir outro professor, acesse a página de "
                                f"<a href='{link_substituicao}' target='_blank' class='alert-link'>Substituições</a>.</small>"
                            )
                            
                            conflitos_encontrados.append(mark_safe(mensagem_html))

                if conflitos_encontrados:
                    if is_ajax:
                        return JsonResponse(
                            {"success": False, "errors": conflitos_encontrados},
                            status=400,
                        )
                    for erro in conflitos_encontrados:
                        messages.error(request, erro)
                else:
                    aulas_criadas_count = 0
                    aula_principal_criada = None
                    for data_agendamento in datas_para_agendar:
                        # Agora, a variável 'status' sempre terá um valor válido.
                        nova_aula = Aula.objects.create(
                            modalidade=modalidade,
                            data_hora=data_agendamento,
                            status=status,
                        )
                        nova_aula.alunos.set(list(alunos_ids))
                        nova_aula.professores.set(list(professores_ids))
                        if aulas_criadas_count == 0:
                            aula_principal_criada = nova_aula
                        aulas_criadas_count += 1

                    if reposicao_de_id_hidden and aula_principal_criada:
                        try:
                            presenca_a_repor = PresencaAluno.objects.get(
                                id=reposicao_de_id_hidden
                            )
                            presenca_a_repor.aula_reposicao = aula_principal_criada
                            presenca_a_repor.save()
                            message_text = f"Aula de reposição para '{presenca_a_repor.aluno.nome_completo}' agendada com sucesso!"
                        except PresencaAluno.DoesNotExist:
                            message_text = "Aula agendada, mas não foi possível vincular à falta original (ID não encontrado)."
                            messages.warning(request, message_text)
                    else:
                        message_text = (
                            f"{aulas_criadas_count} aulas recorrentes foram agendadas."
                            if aulas_criadas_count > 1
                            else "Aula agendada com sucesso!"
                        )

                    if is_ajax:
                        return JsonResponse({"success": True, "message": message_text})

                    messages.success(request, message_text)
                    return redirect("scheduler:dashboard")

    # Passamos o usuário aqui também para o caso de renderização GET
    form = AulaForm(request.POST or None, initial=initial_form_data, user=request.user)
    aluno_formset = AlunoFormSet(
        request.POST or None, initial=initial_aluno_data, prefix="alunos"
    )
    professor_formset = ProfessorFormSet(request.POST or None, prefix="professores")

    contexto = {
        "form": form,
        "aluno_formset": aluno_formset,
        "professor_formset": professor_formset,
        "titulo": "Agendar Nova Aula",
        "form_action": reverse("scheduler:aula_agendar"),
    }

    if presenca_original:
        contexto["presenca_original"] = presenca_original
        contexto["titulo"] = (
            f"Agendar Reposição para {presenca_original.aluno.nome_completo}"
        )
        contexto["form_action"] = (
            f"{reverse('scheduler:aula_agendar')}?reposicao_de={presenca_original.id}"
        )

    return render(request, "scheduler/aula_form.html", contexto)


@login_required
@never_cache
def editar_aula(request, pk):
    aula = get_object_or_404(Aula, pk=pk)

    # Lógica de permissão para edição (sem alterações)
    pode_editar = False
    if request.user.tipo in ["admin", "comercial"]:
        pode_editar = True
    elif request.user.tipo == "professor" and request.user in aula.professores.all():
        pode_editar = True

    if not pode_editar:
        messages.error(request, "Você não tem permissão para editar esta aula.")
        return redirect("scheduler:dashboard")

    AlunoFormSet = formset_factory(AlunoChoiceForm, extra=1, can_delete=True)
    ProfessorFormSet = formset_factory(ProfessorChoiceForm, extra=1, can_delete=True)

    if request.method == "POST":
        form = AulaForm(request.POST, instance=aula, user=request.user)
        aluno_formset = AlunoFormSet(request.POST, prefix="alunos")

        # Validação condicional do formset de professores (sem alterações)
        professor_formset_is_valid = False
        if request.user.tipo in ["admin", "comercial"]:
            professor_formset = ProfessorFormSet(request.POST, prefix="professores")
            if professor_formset.is_valid():
                professor_formset_is_valid = True
        else:
            professor_formset = ProfessorFormSet(prefix="professores")
            professor_formset_is_valid = True

        if form.is_valid() and aluno_formset.is_valid() and professor_formset_is_valid:
            # O formulário agora lida com o fuso horário automaticamente.
            data_hora_nova = form.cleaned_data.get("data_hora")

            is_recorrente = form.cleaned_data.get("recorrente_mensal")

            alunos_ids = {
                f["aluno"].id
                for f in aluno_formset.cleaned_data
                if f and not f.get("DELETE")
            }

            if request.user.tipo in ["admin", "comercial"]:
                professores_ids = {
                    f["professor"].id
                    for f in professor_formset.cleaned_data
                    if f and not f.get("DELETE")
                }
            else:
                professores_ids = {p.id for p in aula.professores.all()}

            # Lógica de recorrência na edição
            modalidade = form.cleaned_data.get("modalidade")
            is_ac = "atividade complementar" in modalidade.nome.lower()

            # 1. Checa conflito de Professor
            conflito_info_prof = _check_conflito_aula(
                list(professores_ids), data_hora_nova, aula_id=aula.pk
            )
            
            # 2. Checa conflito de Aluno
            conflito_info_aluno = {"conflito": False} # Valor padrão
            if not is_ac: # Só checa aluno se não for AC
                conflito_info_aluno = _check_conflito_aluno(
                    list(alunos_ids), data_hora_nova, aula_id=aula.pk
                )

            if conflito_info_prof["conflito"]:
                messages.error(
                    request,
                    f"Não foi possível atualizar a aula: {conflito_info_prof['mensagem']}",
                )
            elif conflito_info_aluno["conflito"]:
                # ★ MENSAGEM RICA E LINK CORRETO APLICADOS AQUI ★
                aluno_nome = conflito_info_aluno.get('aluno_nome', 'Um dos alunos')
                aula_pk = conflito_info_aluno.get('aula_conflitante_pk')
                prof_nomes = conflito_info_aluno.get('professores_conflito', 'N/A')
                
                link_aula = reverse('scheduler:aula_validar', args=[aula_pk])
                # ★ LINK CORRIGIDO ★
                link_substituicao = reverse('scheduler:aulas_para_substituir')
                
                mensagem_html = (
                    f"<b>Conflito:</b> O aluno <strong>{aluno_nome}</strong> já tem uma aula com <strong>{prof_nomes}</strong> neste horário. "
                    f"<a href='{link_aula}' target='_blank' class='alert-link'>Clique para substituir essa aula</a>."
                    f"<br><small>Caso queira substituir outro professor, acesse a página de "
                    f"<a href='{link_substituicao}' target='_blank' class='alert-link'>Substituições</a>.</small>"
                )
                messages.error(request, mark_safe(mensagem_html))
            else:
                aula_salva = form.save(
                    commit=False
                )  # Adicionado commit=False para controle
                aula_salva.alunos.set(list(alunos_ids))
                aula_salva.professores.set(list(professores_ids))
                aula_salva.save()  # Salva o M2M

                if is_recorrente:
                    datas_para_agendar = []
                    dia_semana = data_hora_nova.weekday()
                    mes, ano = data_hora_nova.month, data_hora_nova.year
                    cal = calendar.Calendar()

                    for dia in cal.itermonthdates(ano, mes):
                        if (
                            dia.month == mes
                            and dia.weekday() == dia_semana
                            and dia > data_hora_nova.date()
                        ):
                            nova_data_hora = data_hora_nova.replace(
                                year=dia.year, month=dia.month, day=dia.day
                            )
                            datas_para_agendar.append(nova_data_hora)

                    conflitos_novos = [
                        info["mensagem"]
                        for dt in datas_para_agendar
                        if (info := _check_conflito_aula(list(professores_ids), dt))[
                            "conflito"
                        ]
                    ]

                    if conflitos_novos:
                        messages.warning(
                            request,
                            f"A aula do dia {data_hora_nova.strftime('%d/%m')} foi atualizada, mas as aulas recorrentes não puderam ser criadas devido a conflitos.",
                        )
                        for erro in conflitos_novos:
                            messages.error(request, erro)
                    else:
                        aulas_criadas_count = 0
                        for data_agendamento in datas_para_agendar:
                            nova_aula = Aula.objects.create(
                                modalidade=aula_salva.modalidade,
                                data_hora=data_agendamento,
                                status=aula_salva.status,
                            )
                            nova_aula.alunos.set(list(alunos_ids))
                            nova_aula.professores.set(list(professores_ids))
                            aulas_criadas_count += 1

                        if aulas_criadas_count > 0:
                            messages.success(
                                request,
                                f"Aula principal atualizada e {aulas_criadas_count} novas aulas recorrentes foram agendadas!",
                            )
                        else:
                            messages.success(request, "Aula atualizada com sucesso!")
                else:
                    messages.success(request, "Aula atualizada com sucesso!")

                return redirect("scheduler:dashboard")

    else:  # GET
        form = AulaForm(instance=aula, user=request.user)
        alunos_data = [{"aluno": aluno_obj} for aluno_obj in aula.alunos.all()]
        professores_data = [
            {"professor": prof_obj} for prof_obj in aula.professores.all()
        ]
        aluno_formset = AlunoFormSet(initial=alunos_data, prefix="alunos")
        professor_formset = ProfessorFormSet(
            initial=professores_data, prefix="professores"
        )

    contexto = {
        "form": form,
        "aula": aula,
        "aluno_formset": aluno_formset,
        "professor_formset": professor_formset,
        "titulo": "Editar Aula",
        "form_action": reverse("scheduler:aula_editar", kwargs={"pk": aula.pk}),
    }
    return render(request, "scheduler/aula_form.html", contexto)


@login_required
@require_POST
def excluir_aula(request, pk):
    aula = get_object_or_404(Aula, pk=pk)
    aula.delete()
    messages.success(request, "Aula excluída com sucesso!")
    return redirect("scheduler:dashboard")


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "professor"])
def aulas_para_substituir(request):
    """
    Mostra uma lista de aulas futuras agendadas para outros professores,
    disponíveis para substituição, COM FILTROS AVANÇADOS.
    """

    # --- 1. QUERYSET BASE (LÓGICA ORIGINAL MANTIDA) ---
    aulas_queryset = Aula.objects.filter( status="Agendada").exclude(
        professores=request.user
    )

    # --- 2. LEITURA DOS FILTROS DA URL ---
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get(
        "professor_filtro", ""
    )  # Apenas Admins usarão
    modalidade_filtro_id = request.GET.get("modalidade_filtro", "")
    aluno_filtro_ids = request.GET.getlist("aluno_filtro")

    # --- 3. APLICAÇÃO DOS FILTROS ADICIONAIS ---
    if data_inicial_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__gte=datetime.strptime(
                    data_inicial_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__lte=datetime.strptime(
                    data_final_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if modalidade_filtro_id:
        aulas_queryset = aulas_queryset.filter(modalidade_id=modalidade_filtro_id)

    # Filtro de aluno
    if aluno_filtro_ids:
        aluno_filtro_ids_validos = [int(id) for id in aluno_filtro_ids if id.isdigit()]
        if aluno_filtro_ids_validos:
            aulas_queryset = aulas_queryset.filter(
                alunos__id__in=aluno_filtro_ids_validos
            ).distinct()

    # Filtro de professor (só se aplica se o usuário for admin)
    if request.user.tipo == "admin" and professor_filtro_id:
        aulas_queryset = aulas_queryset.filter(professores__id=professor_filtro_id)

    # --- 4. ORDENAÇÃO E PAGINAÇÃO ---
    aulas_ordenadas = aulas_queryset.distinct().order_by("data_hora")
    paginator = Paginator(aulas_ordenadas, 10)
    aulas_paginadas = paginator.get_page(request.GET.get("page"))

    # --- 5. MONTAGEM DO CONTEXTO COMPLETO PARA O TEMPLATE ---
    contexto = {
        "titulo": "Aulas Disponíveis para Substituição",
        "aulas": aulas_paginadas,  # Para a paginação
        "page_obj": aulas_paginadas,  # Para o componente de paginação
        # --- Dados para o componente de filtro ---
        "professores_list": CustomUser.objects.filter(
            tipo__in=["professor", "admin"]
        ).order_by("username"),
        "modalidades_list": Modalidade.objects.all().order_by("nome"),
        "alunos_list": Aluno.objects.all().order_by("nome_completo"),
        # --- Valores atuais dos filtros para preencher o formulário ---
        "data_inicial": data_inicial_str,
        "data_final": data_final_str,
        "professor_filtro": professor_filtro_id,
        "modalidade_filtro": modalidade_filtro_id,
        "aluno_filtro_ids": [int(pk) for pk in aluno_filtro_ids if pk.isdigit()],
        "request": request,
    }
    return render(request, "scheduler/aulas_para_substituir.html", contexto)


# --- VIEWS DE GERENCIAMENTO DE ALUNOS ---
@login_required
def listar_alunos(request):
    search_query = request.GET.get("q", "")
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    if 'status_filtro' in request.GET:
        status_filtro = request.GET.get("status_filtro", "")
    else:
        status_filtro = 'ativo'
    now = timezone.now()

    if request.user.tipo == "professor":
        alunos_queryset = (
            Aluno.objects.filter(
                Q(aulas_aluno__professores=request.user) | Q(aulas_aluno__isnull=True)
            )
            .distinct()
            .annotate(
                total_aulas=Count(
                    "aulas_aluno", filter=Q(aulas_aluno__professores=request.user)
                ),
                proxima_aula=Min(
                    Case(
                        When(
                            aulas_aluno__data_hora__gte=now,
                            aulas_aluno__professores=request.user,
                            then="aulas_aluno__data_hora",
                        ),
                        default=None,
                    )
                ),
            )
        )
    else:
        alunos_queryset = Aluno.objects.all().annotate(
            total_aulas=Count("aulas_aluno"),
            proxima_aula=Min(
                Case(
                    When(
                        aulas_aluno__data_hora__gte=now, then="aulas_aluno__data_hora"
                    ),
                    default=None,
                )
            ),
        )

    if search_query:
        alunos_queryset = alunos_queryset.filter(
            Q(nome_completo__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(telefone__icontains=search_query)
        ).distinct()

    if status_filtro:
        alunos_queryset = alunos_queryset.filter(status=status_filtro)

    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            alunos_queryset = alunos_queryset.filter(data_criacao__gte=data_inicial)
        except ValueError:
            pass

    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
            alunos_queryset = alunos_queryset.filter(data_criacao__lte=data_final)
        except ValueError:
            pass

    orderby = request.GET.get("orderby", "nome_completo")
    if orderby not in ["nome_completo", "-nome_completo", "proxima_aula", "-proxima_aula"]:
        orderby = "nome_completo"

    alunos_queryset = alunos_queryset.order_by(orderby)

    alunos = alunos_queryset.order_by("nome_completo")
    aulas_ordenadas = alunos.order_by("nome_completo")
    total_alunos_filtrados = alunos.count()
    paginator = Paginator(aulas_ordenadas, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    contexto = {
        "alunos": alunos,
        "titulo": "Gerenciamento de Alunos",
        "total_alunos_filtrados": total_alunos_filtrados,
        "search_query": search_query,
        "data_inicial": data_inicial_str,
        "data_final": data_final_str,
        "status_filtro": status_filtro,
        "status_choices": Aluno.STATUS_CHOICES,
        "orderby": orderby,
        "page_obj": page_obj,
    }
    return render(request, "scheduler/aluno_listar.html", contexto)


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "professor", "comercial"])
def criar_aluno(request):
    if request.method == "POST":
        form = AlunoForm(request.POST)
        if form.is_valid():
            novo_aluno = form.save()

            if form.cleaned_data.get("criar_recorrencia"):
                if novo_aluno.valor_mensalidade and novo_aluno.valor_mensalidade > 0:
                    unidade_ativa_id = request.session.get("unidade_ativa_id")
                    if unidade_ativa_id:
                        categoria_mensalidade, _ = Category.objects.get_or_create(
                            name="Mensalidade",
                            type="income",
                            unidade_negocio_id=unidade_ativa_id,
                            defaults={"tipo_dre": "receita"},
                        )

                        ReceitaRecorrente.objects.create(
                            unidade_negocio_id=unidade_ativa_id,
                            aluno=novo_aluno,
                            categoria=categoria_mensalidade,
                            descricao=f"Recorrência de Mensalidade para {novo_aluno.nome_completo}",
                            ativa=True,
                            data_inicio=novo_aluno.data_criacao
                            or timezone.now().date(),
                        )
                        messages.info(
                            request,
                            "Regra de recorrência de mensalidade criada automaticamente.",
                        )
                    else:
                        messages.warning(
                            request,
                            "A recorrência não foi criada pois nenhuma unidade de negócio está ativa na sessão.",
                        )

            lead_id = request.POST.get("lead_id")
            if lead_id:
                try:
                    lead = Lead.objects.get(id=lead_id)
                    lead.status = "convertido"
                    lead.aluno_convertido = novo_aluno
                    lead.convertido_por = request.user
                    lead.save()
                    messages.info(
                        request,
                        f"Lead '{lead.nome_interessado}' convertido com sucesso!",
                    )
                except Lead.DoesNotExist:
                    pass

            messages.success(request, "Aluno criado com sucesso!")
            return redirect("scheduler:aluno_listar")
        
    else:
        initial_data = {}
        if "lead_id" in request.GET:
            initial_data["nome_completo"] = request.GET.get("nome_completo")
            initial_data["responsavel_nome"] = request.GET.get("responsavel_nome")
            initial_data["email"] = request.GET.get("email")
            initial_data["telefone"] = request.GET.get("telefone")

        form = AlunoForm(initial=initial_data)

    contexto = {"form": form, "titulo": "Adicionar Novo Aluno"}
    if "lead_id" in request.GET:
        contexto["lead_id"] = request.GET.get("lead_id")

    return render(
        request,
        "scheduler/aluno_form.html",
        contexto,
    )


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "professor"])
def editar_aluno(request, pk):
    aluno = get_object_or_404(Aluno, pk=pk)
    if request.method == "POST":
        form = AlunoForm(request.POST, instance=aluno)
        if form.is_valid():
            form.save()
            messages.success(request, "Aluno atualizado com sucesso!")
            return redirect("scheduler:aluno_listar")
    else:
        form = AlunoForm(instance=aluno)
    contexto = {"form": form, "titulo": "Editar Aluno"}
    return render(request, "scheduler/aluno_form.html", contexto)


@user_passes_test(is_admin)
@require_POST
def excluir_aluno(request, pk):
    aluno = get_object_or_404(Aluno, pk=pk)
    aluno.delete()
    messages.success(request, "Aluno excluído com sucesso!")
    return redirect("scheduler:aluno_listar")


def calculate_moving_average(data_points, window_size=3):
    """Calcula a média móvel para uma lista de pontos de dados."""
    if not data_points:
        return []

    # Extrai os valores de 'y' para o cálculo
    values = [p["y"] for p in data_points]
    moving_averages = []

    for i in range(len(values)):
        if i < window_size - 1:
            moving_averages.append(None)  # Não há dados suficientes para a média
        else:
            window = values[i - window_size + 1 : i + 1]
            average = sum(window) / window_size
            moving_averages.append(round(average, 2))

    # Remonta a estrutura com 'x' e 'y'
    result = [
        {"x": data_points[i]["x"], "y": moving_averages[i]}
        for i in range(len(data_points))
    ]
    return result


@login_required
def detalhe_aluno(request, pk):
    aluno = get_object_or_404(Aluno, pk=pk)
    historico_financeiro = aluno.financial_transactions.filter(
        category__type="income"
    ).order_by("-transaction_date")

    # --- INÍCIO DA LÓGICA DE FILTRO GLOBAL ---
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    status_filtro = request.GET.get("status_filtro", "")

    # Query base para todas as aulas do aluno
    aulas_do_aluno = Aula.objects.filter(alunos=aluno)

    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            aulas_do_aluno = aulas_do_aluno.filter(data_hora__date__gte=data_inicial)
        except (ValueError, TypeError):
            pass
    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
            aulas_do_aluno = aulas_do_aluno.filter(data_hora__date__lte=data_final)
        except (ValueError, TypeError):
            pass

    if status_filtro:
        aulas_do_aluno = aulas_do_aluno.filter(status=status_filtro)
    # --- FIM DA LÓGICA DE FILTRO GLOBAL ---

    aulas_do_aluno = aulas_do_aluno.select_related(
        "modalidade", "relatorioaula__professor_que_validou"
    ).prefetch_related("professores", "presencas")

    presenca_status_subquery = PresencaAluno.objects.filter(
        aula=OuterRef("pk"), aluno=aluno
    ).values("status")[:1]
    aulas_com_presenca = aulas_do_aluno.annotate(
        status_presenca_aluno=Subquery(presenca_status_subquery)
    )

    total_realizadas = aulas_com_presenca.filter(
        status__in=["Realizada", "Aluno Ausente"], status_presenca_aluno="presente"
    ).count()
    total_ausencias = (
        aulas_com_presenca.filter(
            status__in=["Realizada", "Aluno Ausente"], status_presenca_aluno="ausente"
        )
        .exclude(status="Reposta")
        .count()
    )
    total_canceladas = aulas_com_presenca.filter(status="Cancelada").count()
    total_agendadas = aulas_com_presenca.filter(status="Agendada").count()
    total_aulas = aulas_do_aluno.count()

    aulas_contabilizadas_para_presenca = total_realizadas + total_ausencias
    taxa_presenca = (
        (total_realizadas / aulas_contabilizadas_para_presenca) * 100
        if aulas_contabilizadas_para_presenca > 0
        else 0
    )

    aulas_presente = aulas_com_presenca.filter(status_presenca_aluno="presente")

    top_professores = (
        aulas_presente.filter(professores__isnull=False)
        .values("professores__pk", "professores__username")
        .annotate(contagem=Count("professores__pk"))
        .order_by("-contagem")[:3]
    )
    top_modalidades = (
        aulas_presente.values("modalidade__nome")
        .annotate(contagem=Count("modalidade"))
        .order_by("-contagem")[:3]
    )

    chart_labels = ["Realizadas", "Ausências", "Canceladas", "Agendadas"]
    chart_data = [
        total_realizadas,
        total_ausencias,
        total_canceladas,
        total_agendadas,
    ]

    # 🔹 Somente datas reais de aulas com prática registrada
    relatorios_de_aulas_presente_ids = aulas_presente.filter(
        relatorioaula__isnull=False
    ).values_list("relatorioaula__pk", flat=True)

    dados_evolucao = (
        ItemRudimento.objects.filter(
            relatorio_id__in=list(relatorios_de_aulas_presente_ids),
            bpm__isnull=False,
        )
        .exclude(bpm__exact="")
        .order_by("relatorio__aula__data_hora")
    )

    dados_grafico_por_exercicio = {}
    lista_exercicios_unicos = []
    evolucao_total_aulas = 0

    if dados_evolucao.exists():
        dados_agrupados = defaultdict(list)

        for item in dados_evolucao:
            try:
                bpm = int(str(item.bpm).strip().replace("bpm", ""))
                item_date = item.relatorio.aula.data_hora.date()
                descricao = item.descricao.strip().title()
                dados_agrupados[descricao].append(
                    {"x": item_date.isoformat(), "y": bpm}
                )
            except (ValueError, AttributeError):
                continue

        for descricao, pontos in dados_agrupados.items():
            pontos_ordenados = sorted(pontos, key=lambda p: p["x"])
            dados_grafico_por_exercicio[descricao] = {
                "data": pontos_ordenados,
                "moving_average": calculate_moving_average(
                    pontos_ordenados, window_size=3
                ),
            }

        lista_exercicios_unicos = sorted(dados_agrupados.keys())
        evolucao_total_aulas = (
            dados_evolucao.values("relatorio__aula__data_hora__date").distinct().count()
        )

    historico_aulas_qs = aulas_com_presenca.order_by("-data_hora")
    paginator = Paginator(historico_aulas_qs, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    for aula in page_obj.object_list:
        if aula.status in ["Realizada", "Aluno Ausente"]:
            presencas_map = {p.aluno_id: p.status for p in aula.presencas.all()}
            aula.alunos_com_status = []
            for a in aula.alunos.all():
                status = presencas_map.get(a.id, "nao_lancado")
                aula.alunos_com_status.append({"aluno": a, "status": status})

    contexto = {
        "aluno": aluno,
        "aulas_do_aluno": page_obj,
        "request": request,
        "titulo": f"Perfil do Aluno: {aluno.nome_completo}",
        "total_aulas": total_aulas,
        "total_realizadas": total_realizadas,
        "total_canceladas": total_canceladas,
        "total_aluno_ausente": total_ausencias,
        "taxa_presenca": taxa_presenca,
        "top_professores": top_professores,
        "top_modalidades": top_modalidades,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "historico_financeiro": historico_financeiro,
        "evolucao_total_aulas": evolucao_total_aulas,
        "dados_grafico_por_exercicio": dados_grafico_por_exercicio,
        "lista_exercicios_unicos": lista_exercicios_unicos,
        # Filtros
        "data_inicial": data_inicial_str,
        "data_final": data_final_str,
        "status_filtro": status_filtro,
    }
    return render(request, "scheduler/aluno_detalhe.html", contexto)


@login_required
def listar_aulas(request):
    # --- 1 a 4: Lógica de filtros permanece a mesma ---
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get("professor_filtro", "")
    modalidade_filtro_id = request.GET.get("modalidade_filtro", "")
    status_filtro = request.GET.get("status_filtro", "")
    aluno_filtro_ids = request.GET.getlist("aluno_filtro")
    if request.user.tipo in ["admin", "comercial"]:
        aulas_queryset = Aula.objects.all()
        contexto_titulo = "Histórico Geral de Aulas"
    else:
        aulas_queryset = Aula.objects.filter(
            Q(professores=request.user)
            | Q(relatorioaula__professor_que_validou=request.user)
        ).distinct()
        contexto_titulo = "Meu Histórico de Aulas"
    if data_inicial_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__gte=datetime.strptime(
                    data_inicial_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__lte=datetime.strptime(
                    data_final_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if modalidade_filtro_id:
        aulas_queryset = aulas_queryset.filter(modalidade_id=modalidade_filtro_id)
    if aluno_filtro_ids:
        aluno_filtro_ids_validos = [int(id) for id in aluno_filtro_ids if id.isdigit()]
        if aluno_filtro_ids_validos:
            aulas_queryset = aulas_queryset.filter(
                alunos__id__in=aluno_filtro_ids_validos
            ).distinct()
    if professor_filtro_id:
        aulas_queryset = aulas_queryset.filter(
            Q(professores__id=professor_filtro_id)
            | Q(relatorioaula__professor_que_validou__id=professor_filtro_id)
        ).distinct()
    if status_filtro:
        if status_filtro == "Substituído":
            aulas_queryset = aulas_queryset.filter(
                status="Realizada",
                relatorioaula__professor_que_validou__isnull=False,
                professores__isnull=False,
            ).exclude(professores=F("relatorioaula__professor_que_validou"))
        elif status_filtro == "professor_ausente":
            aulas_queryset = aulas_queryset.filter(
                modalidade__nome__icontains="atividade complementar",
                presencas_professores__status="ausente",
            )
        else:
            aulas_queryset = aulas_queryset.filter(status=status_filtro)

    # --- 5. PREPARAÇÃO FINAL E PAGINAÇÃO ---
    # ★★★ CORREÇÃO 1: Usando o related_name correto 'presencas'. ★★★
    aulas_ordenadas = (
        aulas_queryset.distinct()
        .order_by("-data_hora")
        .prefetch_related(
            "alunos",
            "professores",
            "relatorioaula__professor_que_validou",
            "presencas_professores__professor",
            "presencas",
        )
    )

    paginator = Paginator(aulas_ordenadas, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    # ★★★ CORREÇÃO 2: Usando 'aula.presencas.all()' para iterar. ★★★
    for aula in page_obj.object_list:
        if aula.status in ["Realizada", "Aluno Ausente"]:
            presencas_map = {p.aluno_id: p.status for p in aula.presencas.all()}
            aula.alunos_com_status = []
            for aluno in aula.alunos.all():
                status = presencas_map.get(aluno.id, "nao_lancado")
                aula.alunos_com_status.append({"aluno": aluno, "status": status})

    # ★★★ CORREÇÃO 3: Linhas duplicadas de paginação removidas. ★★★

    # --- 6. MONTAGEM DO CONTEXTO PARA O TEMPLATE ---
    contexto = {
        "aulas": page_obj,
        "page_obj": page_obj,
        "titulo": contexto_titulo,
        "data_inicial": data_inicial_str,
        "data_final": data_final_str,
        "professor_filtro": professor_filtro_id,
        "modalidade_filtro": modalidade_filtro_id,
        "status_filtro": status_filtro,
        "aluno_filtro_ids": [int(i) for i in aluno_filtro_ids if i.isdigit()],
        "professores_list": CustomUser.objects.filter(
            tipo__in=["professor", "admin"]
        ).order_by("username"),
        "modalidades_list": Modalidade.objects.all().order_by("nome"),
        "alunos_list": Aluno.objects.all().order_by("nome_completo"),
        "status_choices": Aula.STATUS_AULA_CHOICES,
        "request": request,
    }
    return render(request, "scheduler/aula_listar.html", contexto)


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "professor", "comercial"])
@never_cache
def validar_aula(request, pk):
    aula = get_object_or_404(Aula, pk=pk)
    relatorio, created = RelatorioAula.objects.get_or_create(aula=aula)

    pode_editar = False
    
    # --- LÓGICA DE PERMISSÃO DE EDIÇÃO ---
    if request.user.tipo == "admin":
        pode_editar = True
    elif request.user.tipo == "professor":
        if aula.status == "Agendada" and request.user in aula.professores.all():
            pode_editar = True
        elif relatorio.professor_que_validou == request.user:
            pode_editar = True
        elif not relatorio.professor_que_validou:
            pode_editar = True
    # REGRA COMERCIAL: Só edita se estiver na lista de professores da aula
    elif request.user.tipo == "comercial":
        if request.user in aula.professores.all():
            pode_editar = True

    # Define o modo de visualização inicial
    view_mode = "visualizar"
    if pode_editar:
        if aula.status == "Agendada" or request.GET.get("mode") == "editar":
            view_mode = "editar"

    is_ac = "atividade complementar" in aula.modalidade.nome.lower()

    if is_ac:
        professores_da_aula = aula.professores.all()
        for prof in professores_da_aula:
            PresencaProfessor.objects.get_or_create(aula=aula, professor=prof)
        presenca_queryset = PresencaProfessor.objects.filter(aula=aula).order_by(
            "professor__username"
        )
        presenca_formset_class = PresencaProfessorFormSet
        presenca_prefix = "presencas_prof"
    else:
        alunos_da_aula = aula.alunos.all()
        if alunos_da_aula.exists():
            for aluno in alunos_da_aula:
                PresencaAluno.objects.get_or_create(aula=aula, aluno=aluno)
        presenca_queryset = PresencaAluno.objects.filter(aula=aula).order_by(
            "aluno__nome_completo"
        )
        presenca_formset_class = PresencaAlunoFormSet
        presenca_prefix = "presencas_alunos"

    historico_ultima_aula = None
    # (A sua lógica de buscar histórico de última aula permanece aqui, sem alterações)
    if is_ac:
        ultima_aula = (
            Aula.objects.filter(
                modalidade=aula.modalidade,
                status="Realizada",
                data_hora__lt=aula.data_hora,
                relatorioaula__isnull=False,
            )
            .order_by("-data_hora")
            .first()
        )
    else:
        if "alunos_da_aula" in locals() and alunos_da_aula.exists():
            query_aula_relevante = (
                Aula.objects.filter(
                    alunos__in=alunos_da_aula,
                    status="Realizada",
                    data_hora__lt=aula.data_hora,
                    relatorioaula__isnull=False,
                )
                .filter(
                    Q(relatorioaula__itens_rudimentos__isnull=False)
                    | Q(relatorioaula__itens_ritmo__isnull=False)
                    | Q(relatorioaula__itens_viradas__isnull=False)
                    | ~Q(relatorioaula__conteudo_teorico__exact="")
                    | ~Q(relatorioaula__repertorio_musicas__exact="")
                )
                .distinct()
                .order_by("-data_hora")
                .first()
            )
            ultima_aula = query_aula_relevante
        else:
            ultima_aula = None
    if ultima_aula:
        historico_ultima_aula = ultima_aula.relatorioaula

    if request.method == "POST":
        if not pode_editar:
            messages.error(request, "Você só pode validar ou editar relatórios de aulas onde você é o professor responsável.")
            return redirect("scheduler:aula_validar", pk=aula.pk)

        form = RelatorioAulaForm(request.POST, instance=relatorio)
        presenca_formset = presenca_formset_class(
            request.POST, queryset=presenca_queryset, prefix=presenca_prefix
        )
        rudimentos_formset = ItemRudimentoFormSet(
            request.POST, instance=relatorio, prefix="rudimentos"
        )
        ritmo_formset = ItemRitmoFormSet(
            request.POST, instance=relatorio, prefix="ritmo"
        )
        viradas_formset = ItemViradaFormSet(
            request.POST, instance=relatorio, prefix="viradas"
        )

        form_valid = form.is_valid()
        presenca_valid = presenca_formset.is_valid()
        rudimentos_valid = rudimentos_formset.is_valid()
        ritmo_valid = ritmo_formset.is_valid()
        viradas_valid = viradas_formset.is_valid()

        if not form_valid:
            print("Form principal inválido:", form.errors)
            messages.warning(
                request,
                "Ops! Ocorreu um erro ao salvar o relatório. Por favor, revise os campos."
            )

        if not presenca_valid:
            print("Formset de presença inválido:", presenca_formset.errors)
            messages.warning(
                request,
                "Atenção: Não foi possível salvar a lista de presença. Por favor, verifique se o status (Presente/Ausente) de cada participante foi selecionado corretamente."
            )

        if not rudimentos_valid:
            print("Formset de rudimentos inválido:", rudimentos_formset.errors)
            messages.warning(
                request,
                "Ops! Parece que na seção de Rudimentos, uma linha de exercício foi deixada em branco.\nPara salvar, preencha a descrição do exercício ou clique na lixeira para remover a linha vazia."
            )

        if not ritmo_valid:
            print("Formset de ritmo inválido:", ritmo_formset.errors)
            messages.warning(
                request,
                "Ops! Parece que na seção de Ritmos, uma linha de exercício foi deixada em branco. Para salvar, preencha a descrição do exercício ou clique na lixeira para remover a linha vazia."
            )

        if not viradas_valid:
            print("Formset de viradas inválido:", viradas_formset.errors)
            messages.warning(
                request,
                "Atenção na seção de Viradas: você deixou uma linha de exercício sem descrição. É preciso preenchê-la ou remover a linha clicando no ícone da lixeira antes de salvar."
            )

        if (
            form.is_valid()
            and presenca_formset.is_valid()
            and rudimentos_formset.is_valid()
            and ritmo_formset.is_valid()
            and viradas_formset.is_valid()
        ):
            try:
                with transaction.atomic():
                    if not relatorio.professor_que_validou:
                        if request.user.tipo == 'professor':
                            relatorio.professor_que_validou = request.user

                    elif request.user.tipo == 'admin' and relatorio.professor_que_validou != request.user:
                        relatorio.ultimo_editor = request.user

                    elif request.user == relatorio.professor_que_validou:
                        relatorio.ultimo_editor = request.user

                    relatorio.save()
                    presenca_formset.save()

                    if is_ac:
                        aula.status = "Realizada"
                    else:
                        num_presentes = PresencaAluno.objects.filter(
                            aula=aula, status="presente"
                        ).count()
                        if aula.alunos.exists() and num_presentes == 0:
                            aula.status = "Aluno Ausente"
                        else:
                            aula.status = "Realizada"

                    aula.save()

                    campos_principais_preenchidos = any(
                        [
                            form.cleaned_data.get("conteudo_teorico", "").strip(),
                            form.cleaned_data.get("observacoes_teoria", "").strip(),
                            form.cleaned_data.get("repertorio_musicas", "").strip(),
                            form.cleaned_data.get("observacoes_repertorio", "").strip(),
                            form.cleaned_data.get("observacoes_gerais", "").strip(),
                        ]
                    )
                    itens_adicionados = any(
                        [
                            rudimentos_formset.has_changed(),
                            ritmo_formset.has_changed(),
                            viradas_formset.has_changed(),
                        ]
                    )

                    if campos_principais_preenchidos or itens_adicionados:
                        rudimentos_formset.save()
                        ritmo_formset.save()
                        viradas_formset.save()

                    if aula.status == "Realizada" and hasattr(aula, "aula_reposta_de"):
                        presenca_original = aula.aula_reposta_de
                        aula_original = presenca_original.aula
                        aula_original.status = "Reposta"
                        aula_original.save()
                        messages.info(
                            request,
                            f"A aula original de {aula_original.data_hora.strftime('%d/%m')} foi marcada como 'Reposta'.",
                        )

                messages.success(request, "Relatório da aula salvo com sucesso!")
                return redirect("scheduler:aula_validar", pk=aula.pk)

            except (ValidationError, DatabaseError, Exception) as e:
                import logging

                logger = logging.getLogger(__name__)
                logger.error(f"Erro ao salvar relatório da aula {aula.pk}: {e}")
                logger.error(f"Erro ao salvar dados: {e}", exc_info=True)
                messages.error(request, f"Ocorreu um erro inesperado: {str(e)}")
                messages.error(
                    request,
                    f"Ocorreu um erro inesperado ao salvar. Nenhuma alteração foi feita. Erro: {e}",
                )

    else:
        form = RelatorioAulaForm(instance=relatorio)
        presenca_formset = presenca_formset_class(
            queryset=presenca_queryset, prefix=presenca_prefix
        )
        rudimentos_formset = ItemRudimentoFormSet(
            instance=relatorio, prefix="rudimentos"
        )
        ritmo_formset = ItemRitmoFormSet(instance=relatorio, prefix="ritmo")
        viradas_formset = ItemViradaFormSet(instance=relatorio, prefix="viradas")

    contexto = {
        "aula": aula,
        "is_ac": is_ac,
        "relatorio": relatorio,
        "form": form,
        "presenca_formset": presenca_formset,
        "rudimentos_formset": rudimentos_formset,
        "ritmo_formset": ritmo_formset,
        "viradas_formset": viradas_formset,
        "historico_ultima_aula": historico_ultima_aula,
        "view_mode": view_mode,
        "pode_editar": pode_editar,
    }
    return render(request, "scheduler/aula_validar.html", contexto)


# --- VIEWS PARA GERENCIAMENTO DE MODALIDADES ---
@user_passes_test(is_admin)
def listar_modalidades(request):
    search_query = request.GET.get("q", "")
    now = timezone.now()
    modalidades_queryset = Modalidade.objects.all()

    # ★★★ INÍCIO DA CORREÇÃO ★★★
    # Removemos o filtro de data (Q(aula__data_hora__gte=now)) da contagem de alunos_ativos.
    modalidades_queryset = modalidades_queryset.annotate(
        total_aulas=Count("aula", distinct=True),
        alunos_ativos=Count("aula__alunos", distinct=True),  # <-- CORRIGIDO
        proxima_aula=Min(
            Case(When(aula__data_hora__gte=now, then="aula__data_hora"), default=None)
        ),
    )
    # ★★★ FIM DA CORREÇÃO ★★★

    if search_query:
        modalidades_queryset = modalidades_queryset.filter(nome__icontains=search_query)

    modalidades = modalidades_queryset.order_by("nome")

    contexto = {
        "modalidades": modalidades,
        "titulo": "Gerenciamento de Categorias",
        "search_query": search_query,
    }
    return render(request, "scheduler/modalidade_listar.html", contexto)


@user_passes_test(is_admin)
def criar_modalidade(request):
    if request.method == "POST":
        form = ModalidadeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria criada com sucesso!")
            return redirect("scheduler:modalidade_listar")
    else:
        form = ModalidadeForm()
    contexto = {"form": form, "titulo": "Adicionar Nova Categoria"}
    return render(request, "scheduler/modalidade_form.html", contexto)


@user_passes_test(is_admin)
def editar_modalidade(request, pk):
    modalidade = get_object_or_404(Modalidade, pk=pk)
    if request.method == "POST":
        form = ModalidadeForm(request.POST, instance=modalidade)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria atualizada com sucesso!")
            return redirect("scheduler:modalidade_listar")
    else:
        form = ModalidadeForm(instance=modalidade)
    contexto = {"form": form, "titulo": f"Editar Categoria: {modalidade.nome.title()}"}
    return render(request, "scheduler/modalidade_form.html", contexto)


@user_passes_test(is_admin)
@require_POST
def excluir_modalidade(request, pk):
    modalidade = get_object_or_404(Modalidade, pk=pk)

    # A lógica de proteção permanece, excelente!
    if modalidade.aula_set.exists():
        messages.error(
            request,
            f'Não é possível excluir a categoria "{modalidade.nome}" porque há aulas associadas a ela.',
        )
        return redirect("scheduler:modalidade_listar")

    modalidade.delete()
    messages.success(request, "Categoria excluída com sucesso!")
    return redirect("scheduler:modalidade_listar")


@user_passes_test(is_admin)
def detalhe_modalidade(request, pk):
    modalidade = get_object_or_404(Modalidade, pk=pk)
    aulas_da_modalidade = Aula.objects.filter(modalidade=modalidade)

    # KPIs de Contagem (sem alterações)
    total_aulas = aulas_da_modalidade.count()
    total_realizadas = aulas_da_modalidade.filter(status="Realizada").count()
    alunos_ativos_count = (
        aulas_da_modalidade.filter(alunos__isnull=False)
        .values("alunos")
        .distinct()
        .count()
    )

    # ★★★ INÍCIO DA LÓGICA ATUALIZADA E CORRIGIDA ★★★
    # Vamos usar a mesma lógica "fonte da verdade" da página do professor,
    # mas aplicada a cada professor que já lecionou nesta categoria.

    professores_da_modalidade = (
        CustomUser.objects.annotate(
            # 1. Conta aulas normais que o professor VALIDOU nesta categoria
            normal_count=Count(
                "aulas_validadas_por_mim",
                distinct=True,
                filter=Q(aulas_validadas_por_mim__aula__modalidade=modalidade)
                & Q(aulas_validadas_por_mim__aula__status="Realizada")
                & ~Q(
                    aulas_validadas_por_mim__aula__modalidade__nome__icontains="atividade complementar"
                ),
            ),
            # 2. Conta ACs em que o professor esteve PRESENTE nesta categoria
            ac_count=Count(
                "presencas_registradas",
                distinct=True,
                filter=Q(presencas_registradas__aula__modalidade=modalidade)
                & Q(presencas_registradas__aula__status="Realizada")
                & Q(presencas_registradas__status="presente")
                & Q(
                    presencas_registradas__aula__modalidade__nome__icontains="atividade complementar"
                ),
            ),
        )
        .annotate(
            # 3. Soma as duas contagens para obter o total real
            aulas_na_modalidade_count=F("normal_count")
            + F("ac_count")
        )
        .filter(
            # 4. Mostra apenas os professores que de fato deram aula nesta categoria
            aulas_na_modalidade_count__gt=0
        )
        .order_by("-aulas_na_modalidade_count", "username")
    )

    # O KPI de contagem de professores também usa este resultado para ser consistente
    professores_count = professores_da_modalidade.count()
    # ★★★ FIM DA LÓGICA ATUALIZADA E CORRIGIDA ★★★

    # Dados do Gráfico (sem alterações)
    aulas_por_mes = (
        aulas_da_modalidade.annotate(mes=TruncMonth("data_hora"))
        .values("mes")
        .annotate(contagem=Count("id"))
        .order_by("mes")
    )

    chart_labels = [item["mes"].strftime("%b/%Y") for item in aulas_por_mes]
    chart_data = [item["contagem"] for item in aulas_por_mes]

    # Paginação (sem alterações)
    aulas_paginadas_qs = aulas_da_modalidade.order_by("-data_hora").prefetch_related(
        "alunos", "professores", "presencas"
    )
    aulas_paginadas = Paginator(aulas_paginadas_qs, 10).get_page(
        request.GET.get("page")
    )

    for aula in aulas_paginadas.object_list:
        if aula.status in ["Realizada", "Aluno Ausente"]:
            presencas_map = {p.aluno_id: p.status for p in aula.presencas.all()}
            aula.alunos_com_status = []
            for aluno in aula.alunos.all():
                status = presencas_map.get(aluno.id, "nao_lancado")
                aula.alunos_com_status.append({"aluno": aluno, "status": status})

    contexto = {
        "modalidade": modalidade,
        "titulo": f"Dashboard da Categoria: {modalidade.nome}",
        "total_aulas": total_aulas,
        "total_realizadas": total_realizadas,
        "alunos_ativos_count": alunos_ativos_count,
        "professores_count": professores_count,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "professores_da_modalidade": professores_da_modalidade,
        "aulas": aulas_paginadas,
    }

    return render(request, "scheduler/modalidade_detalhe.html", contexto)


# --- VIEWS PARA GERENCIAMENTO DE PROFESSORES ---
@user_passes_test(is_admin)
def listar_professores(request):
    search_query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "ativos") # Padrão: ativos
    now = timezone.now()

    professores_queryset = CustomUser.objects.filter(tipo__in=["professor", "admin", "comercial"])

    # --- FILTRO DE STATUS (Ativo/Inativo) ---
    if status_filter == "ativos":
        professores_queryset = professores_queryset.filter(is_active=True)
    elif status_filter == "inativos":
        professores_queryset = professores_queryset.filter(is_active=False)

    if search_query:
        professores_queryset = professores_queryset.filter(
            Q(username__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
        )

    # Ordenação e Anotações (mantendo sua lógica atual)
    professores_queryset = professores_queryset.annotate(
        realizadas_normal=Count(
            "aulas_validadas_por_mim",
            distinct=True,
            filter=Q(aulas_validadas_por_mim__aula__status="Realizada")
            & ~Q(aulas_validadas_por_mim__aula__modalidade__nome__icontains="atividade complementar"),
        ),
        realizadas_ac=Count(
            "presencas_registradas",
            distinct=True,
            filter=Q(
                presencas_registradas__aula__status="Realizada",
                presencas_registradas__status="presente",
                presencas_registradas__aula__modalidade__nome__icontains="atividade complementar",
            ),
        ),
        total_alunos_atendidos=Count("aulas_lecionadas__alunos", distinct=True),
        proxima_aula=Subquery(
            Aula.objects.filter(professores=OuterRef("pk"), data_hora__gte=now)
            .order_by("data_hora")
            .values("data_hora")[:1]
        ),
    ).order_by("first_name", "last_name").distinct()

    professores_list = []
    for p in professores_queryset:
        p.total_aulas_realizadas = p.realizadas_normal + p.realizadas_ac
        professores_list.append(p)

    contexto = {
        "professores": professores_list,
        "titulo": "Gerenciamento de Colaboradores",
        "search_query": search_query,
        "status_filter": status_filter, # Enviando para o template
    }
    return render(request, "scheduler/professor_listar.html", contexto)


@user_passes_test(is_admin)
def editar_professor(request, pk):
    professor = get_object_or_404(CustomUser, pk=pk, tipo__in=["admin", "professor", "comercial"])
    if request.method == "POST":
        form = ProfessorForm(request.POST, instance=professor)
        form.fields.pop("password", None)
        form.fields.pop("password_confirm", None)

        if form.is_valid():
            form.save()
            messages.success(
                request, f"Professor {professor.username} atualizado com sucesso!"
            )
            return redirect("scheduler:professor_listar")
    else:
        form = ProfessorForm(instance=professor)
        form.fields.pop("password", None)
        form.fields.pop("password_confirm", None)
        # form.fields['tipo'].widget.attrs['disabled'] = True

    contexto = {"form": form, "titulo": f"Editar Colaborador: {professor.username}"}
    return render(request, "scheduler/professor_form.html", contexto)


@user_passes_test(is_admin)
@require_POST
def excluir_professor(request, pk):
    professor = get_object_or_404(CustomUser, pk=pk)

    if request.user.pk == pk:
        messages.error(request, "Você não pode excluir seu próprio usuário.")
        return redirect("scheduler:professor_listar")

    # A lógica de aviso permanece, o que é ótimo!
    if professor.aulas_lecionadas.exists():
        messages.warning(
            request,
            f'O professor "{professor.username}" foi excluído, mas estava atribuído a {professor.aulas_lecionadas.count()} aulas. Essas aulas agora estão sem professor.',
        )

    professor.delete()
    messages.success(request, "Professor excluído com sucesso!")
    return redirect("scheduler:professor_listar")


@login_required
def detalhe_professor(request, pk):
    # 1. Verificação de Permissão
    if not (
        request.user.tipo == "admin"
        or (request.user.pk == pk and request.user.tipo in ["professor", "comercial"])
    ):
        messages.error(request, "Você não tem permissão para acessar este perfil.")
        return redirect("scheduler:dashboard")

    user_obj = get_object_or_404(CustomUser, pk=pk, tipo__in=["professor", "admin", "comercial"])
    professor = user_obj
    
    # === LÓGICA ACADÊMICA (COMUM A TODOS) ===
    # Processa datas, filtros e KPIs de aulas para popular a aba "Acadêmico"
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    status_filtro = request.GET.get("status_filtro", "")
    status_filtro_display = status_filtro.replace("_", " ") if status_filtro else ""

    data_inicial, data_final = None, None
    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
        except ValueError: pass
    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
        except ValueError: pass

    # Query Base de Aulas
    aulas_relacionadas_base = Aula.objects.filter(
        Q(professores=professor) | Q(relatorioaula__professor_que_validou=professor)
    ).distinct()

    aulas_kpi = aulas_relacionadas_base
    if data_inicial:
        aulas_kpi = aulas_kpi.filter(data_hora__date__gte=data_inicial)
    if data_final:
        aulas_kpi = aulas_kpi.filter(data_hora__date__lte=data_final)

    # Regras de Negócio para KPIs
    q_realizadas_normal = Q(status="Realizada", relatorioaula__professor_que_validou=professor) & ~Q(modalidade__nome__icontains="atividade complementar")
    q_realizadas_ac = Q(status="Realizada", modalidade__nome__icontains="atividade complementar", presencas_professores__professor=professor, presencas_professores__status="presente")
    
    aulas_realizadas_pelo_professor_no_periodo = aulas_kpi.filter(q_realizadas_normal | q_realizadas_ac).distinct()

    # Cálculo dos KPIs Acadêmicos
    total_realizadas = aulas_realizadas_pelo_professor_no_periodo.count()
    alunos_atendidos = PresencaAluno.objects.filter(aula__in=aulas_realizadas_pelo_professor_no_periodo, status="presente").count()
    total_ausencias_professor = aulas_kpi.filter(modalidade__nome__icontains="atividade complementar", presencas_professores__professor=professor, presencas_professores__status="ausente").count()
    total_agendadas = aulas_kpi.filter(status="Agendada", professores=professor).count()
    total_canceladas = aulas_kpi.filter(status="Cancelada", professores=professor).count()
    total_aluno_ausente = aulas_kpi.filter(status="Aluno Ausente", relatorioaula__professor_que_validou=professor).count()
    total_substituido = aulas_kpi.filter(status="Realizada", professores=professor).exclude(relatorioaula__professor_que_validou=professor).exclude(modalidade__nome__icontains="atividade complementar").count()

    taxa_presenca = ((total_realizadas / (total_realizadas + total_aluno_ausente) * 100) if (total_realizadas + total_aluno_ausente) > 0 else 0)

    # Filtragem e Paginação da Tabela Acadêmica
    aulas_para_tabela = aulas_kpi
    if status_filtro:
        if status_filtro == "Realizada": aulas_para_tabela = aulas_para_tabela.filter(q_realizadas_normal | q_realizadas_ac)
        elif status_filtro == "Substituído": aulas_para_tabela = aulas_para_tabela.filter(status="Realizada", professores=professor).exclude(relatorioaula__professor_que_validou=professor).exclude(modalidade__nome__icontains="atividade complementar")
        elif status_filtro == "Aluno Ausente": aulas_para_tabela = aulas_para_tabela.filter(status="Aluno Ausente", relatorioaula__professor_que_validou=professor)
        elif status_filtro in ["Agendada", "Cancelada"]: aulas_para_tabela = aulas_para_tabela.filter(status=status_filtro, professores=professor)

    aulas_para_tabela = aulas_para_tabela.prefetch_related("presencas", "alunos")
    paginator = Paginator(aulas_para_tabela.order_by("-data_hora"), 10)
    aulas_paginadas = paginator.get_page(request.GET.get("page"))

    aulas_por_categoria = aulas_realizadas_pelo_professor_no_periodo.values("modalidade__nome").annotate(contagem=Count("id", distinct=True)).order_by("-contagem")

    # Contexto Base (Acadêmico)
    contexto = {
        'professor': professor,
        'titulo': f"Perfil de {professor.username.title()}",
        # Dados Acadêmicos
        'aulas_do_professor': aulas_paginadas,
        'total_realizadas': total_realizadas,
        'alunos_atendidos': alunos_atendidos,
        'total_agendadas': total_agendadas,
        'total_canceladas': total_canceladas,
        'total_aluno_ausente': total_aluno_ausente,
        'total_substituido': total_substituido,
        'total_ausencias_professor': total_ausencias_professor,
        'taxa_presenca': taxa_presenca,
        'chart_labels': ["Realizadas", "Agendadas", "Ausências Alunos", "Canceladas", "Substituído"],
        'chart_data': [total_realizadas, total_agendadas, total_aluno_ausente, total_canceladas, total_substituido],
        'status_chart_data': [total_realizadas, total_agendadas, total_aluno_ausente, total_canceladas, total_substituido],
        'aulas_por_categoria': aulas_por_categoria,
        'top_alunos': aulas_kpi.filter(alunos__isnull=False).values("alunos__pk", "alunos__nome_completo").annotate(contagem=Count("alunos__pk")).order_by("-contagem")[:3],
        'top_modalidades': aulas_kpi.filter(professores=professor).values("modalidade__nome").annotate(contagem=Count("modalidade")).order_by("-contagem")[:3],
        'data_inicial': data_inicial_str,
        'data_final': data_final_str,
        'status_filtro': status_filtro,
        'status_filtro_display': status_filtro_display,
    }

    # === LÓGICA ESPECÍFICA COMERCIAL (Se for Comercial) ===
    if user_obj.tipo == 'comercial':
        leads_criados_qs = Lead.objects.filter(criado_por=user_obj)
        leads_convertidos_qs = Lead.objects.filter(convertido_por=user_obj)
        interacoes_do_usuario_qs = InteracaoLead.objects.filter(responsavel=user_obj)

        leads_criados_count = leads_criados_qs.count()
        leads_convertidos_count = leads_convertidos_qs.count()
        leads_perdidos_count = leads_criados_qs.filter(status='perdido').count()
        
        taxa_conversao = (leads_convertidos_count / leads_criados_count * 100) if leads_criados_count > 0 else 0
        
        leads_por_fonte_qs = leads_criados_qs.annotate(fonte_limpa=Coalesce('fonte', V('Não Informada'))).values('fonte_limpa').annotate(contagem=Count('id')).order_by('-contagem')[:5]
        conversoes_por_curso_qs = leads_convertidos_qs.filter(curso_interesse__isnull=False).values('curso_interesse').annotate(contagem=Count('id')).order_by('-contagem')
        
        curso_dict = dict(Lead.CURSO_CHOICES)

        # Adiciona dados comerciais ao contexto
        contexto.update({
            'leads_criados': leads_criados_count,
            'leads_convertidos': leads_convertidos_count,
            'leads_perdidos': leads_perdidos_count,
            'total_interacoes': interacoes_do_usuario_qs.count(),
            'leads_contatados_distintos': interacoes_do_usuario_qs.values('lead').distinct().count(),
            'taxa_conversao': f"{taxa_conversao:.1f}",
            'pipeline_novo': leads_criados_qs.filter(status='novo').count(),
            'pipeline_em_contato': leads_criados_qs.filter(status='em_contato').count(),
            'pipeline_negociando': leads_criados_qs.filter(status='negociando').count(),
            'leads_ativos': leads_criados_qs.filter(status__in=['novo', 'em_contato', 'negociando']).order_by('-data_criacao')[:10],
            'ultimas_interacoes': interacoes_do_usuario_qs.order_by('-data_interacao')[:10],
            'fonte_labels': [item['fonte_limpa'].title() for item in leads_por_fonte_qs],
            'fonte_data': [item['contagem'] for item in leads_por_fonte_qs],
            'curso_labels': [curso_dict.get(item['curso_interesse'], 'Outro').title() for item in conversoes_por_curso_qs],
            'curso_data': [item['contagem'] for item in conversoes_por_curso_qs],
        })
        return render(request, 'scheduler/comercial_detalhe.html', contexto)

    # Se for Professor/Admin, usa o template padrão
    return render(request, 'scheduler/professor_detalhe.html', contexto)


@login_required
def filtrar_aulas_professor_ajax(request, pk):
    # Lógica de permissão (sem alterações)
    if not (
        request.user.tipo == "admin"
        or (request.user.pk == pk and request.user.tipo == "professor")
    ):
        return HttpResponse("Acesso negado.", status=403)

    professor = get_object_or_404(CustomUser, pk=pk, tipo__in=["professor", "admin"])

    # Pega os parâmetros da requisição AJAX (sem alterações)
    status_filtro = request.GET.get("status", "")
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")

    # Queryset base (sem alterações)
    aulas_relacionadas = Aula.objects.filter(
        Q(professores=professor) | Q(relatorioaula__professor_que_validou=professor)
    ).distinct()

    # Aplica filtros de data (sem alterações)
    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            aulas_relacionadas = aulas_relacionadas.filter(
                data_hora__date__gte=data_inicial
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
            aulas_relacionadas = aulas_relacionadas.filter(
                data_hora__date__lte=data_final
            )
        except ValueError:
            pass

    # LÓGICA DE FILTRO POR STATUS (sem alterações)
    if status_filtro:
        if status_filtro == "Realizada":
            q_realizadas_normal = Q(
                status="Realizada", relatorioaula__professor_que_validou=professor
            ) & ~Q(modalidade__nome__icontains="atividade complementar")
            q_realizadas_ac = Q(
                status="Realizada",
                modalidade__nome__icontains="atividade complementar",
                presencas_professores__professor=professor,
                presencas_professores__status="presente",
            )
            aulas_relacionadas = aulas_relacionadas.filter(
                q_realizadas_normal | q_realizadas_ac
            )
        elif status_filtro == "Substituído":
            aulas_relacionadas = (
                aulas_relacionadas.filter(status="Realizada", professores=professor)
                .exclude(relatorioaula__professor_que_validou=professor)
                .exclude(modalidade__nome__icontains="atividade complementar")
            )
        elif status_filtro == "Aluno Ausente":
            aulas_relacionadas = aulas_relacionadas.filter(status="Aluno Ausente")
        else:
            aulas_relacionadas = aulas_relacionadas.filter(
                status=status_filtro, professores=professor
            )

    # ★★★ INÍCIO DA ALTERAÇÃO (1/2) ★★★
    # Adicionamos o prefetch_related para otimizar a busca dos dados de presença dos alunos.
    aulas_relacionadas = aulas_relacionadas.prefetch_related("presencas", "alunos")
    # ★★★ FIM DA ALTERAÇÃO (1/2) ★★★

    # Paginação dos resultados filtrados
    paginator = Paginator(aulas_relacionadas.order_by("-data_hora"), 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    # ★★★ INÍCIO DA ALTERAÇÃO (2/2) ★★★
    # Adicionamos a mesma lógica que processa o status de cada aluno para cada aula.
    for aula in page_obj.object_list:
        if aula.status in ["Realizada", "Aluno Ausente"]:
            presencas_map = {p.aluno_id: p.status for p in aula.presencas.all()}
            aula.alunos_com_status = []
            for aluno in aula.alunos.all():
                status = presencas_map.get(aluno.id, "nao_lancado")
                aula.alunos_com_status.append({"aluno": aluno, "status": status})
    # ★★★ FIM DA ALTERAÇÃO (2/2) ★★★

    # Monta o contexto para o template parcial
    contexto = {
        "aulas_do_professor": page_obj,
        "professor": professor,
        # Passamos os filtros para que a paginação funcione corretamente
        "request": request,
        "status_filtro_ativo": status_filtro,
    }

    # Renderiza APENAS a porção da página que contém a tabela e a paginação
    return render(
        request, "scheduler/partials/professor_detalhe_table_content.html", contexto
    )


# --- NOVA VIEW PARA VERIFICAÇÃO DE CONFLITO VIA AJAX ---
@login_required
def verificar_conflito_aula(request):
    if (
        request.method == "GET"
        and request.headers.get("x-requested-with") == "XMLHttpRequest"
    ):
        professor_id = request.GET.get("professor_id")
        data_hora_str = request.GET.get("data_hora")
        aula_id = request.GET.get("aula_id")  # ID da aula sendo editada (se houver)

        if not professor_id or not data_hora_str:
            return JsonResponse(
                {"conflito": True, "mensagem": "Dados incompletos para verificação."},
                status=400,
            )

        try:
            professor_id = int(professor_id)
            data_hora = datetime.fromisoformat(
                data_hora_str
            )  # Converte a string ISO (YYYY-MM-DDTHH:MM) para datetime
            aula_id = int(aula_id) if aula_id else None  # Converte para int ou None
        except (ValueError, TypeError):
            return JsonResponse(
                {"conflito": True, "mensagem": "Formato de dados inválido."}, status=400
            )

        conflito_info = _check_conflito_aula(professor_id, data_hora, aula_id)
        return JsonResponse(conflito_info)

    return JsonResponse(
        {"conflito": True, "mensagem": "Requisição inválida."}, status=400
    )  # Requisição não-AJAX ou não-GET


def _get_dados_relatorio_agregado(request):
    """
    Função auxiliar que filtra e agrega os dados do relatório.
    """
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get("professor_filtro", "")
    modalidade_filtro_id = request.GET.get("modalidade_filtro", "")
    status_filtro = request.GET.get("status_filtro", "")

    aulas_base_queryset = Aula.objects.all()

    # Aplica filtros
    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            aulas_base_queryset = aulas_base_queryset.filter(
                data_hora__date__gte=data_inicial
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
            aulas_base_queryset = aulas_base_queryset.filter(
                data_hora__date__lte=data_final
            )
        except ValueError:
            pass

    # --- CORRIGIDO ---
    if professor_filtro_id:
        aulas_base_queryset = aulas_base_queryset.filter(
            Q(professores__id=professor_filtro_id)
            | Q(relatorioaula__professor_que_validou__id=professor_filtro_id)
        )
    if modalidade_filtro_id:
        aulas_base_queryset = aulas_base_queryset.filter(
            modalidade_id=modalidade_filtro_id
        )
    if status_filtro:
        aulas_base_queryset = aulas_base_queryset.filter(status=status_filtro)

    aulas_por_professor_final = list(
        aulas_base_queryset.filter(relatorioaula__professor_que_validou__isnull=False)
        .values("relatorioaula__professor_que_validou__username")
        .annotate(aulas_realizadas=Count("id"))
        .order_by("-aulas_realizadas")
        .values("relatorioaula__professor_que_validou__username", "aulas_realizadas")
    )
    aulas_por_modalidade_final = list(
        aulas_base_queryset.values("modalidade__nome", "modalidade__id")
        .annotate(
            total_aulas=Count("id"),
            aulas_realizadas=Count("id", filter=Q(status="Realizada")),
        )
        .order_by("-total_aulas")
        .values("modalidade__id", "modalidade__nome", "total_aulas", "aulas_realizadas")
    )

    return {
        "aulas_por_professor": aulas_por_professor_final,
        "aulas_por_modalidade": aulas_por_modalidade_final,
    }


@user_passes_test(lambda u: u.tipo == "admin")
def relatorios_aulas(request):
    # --- 1. FILTROS ---
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get("professor_filtro")
    modalidade_filtro_id = request.GET.get("modalidade_filtro")
    status_filtro = request.GET.get("status_filtro")
    aluno_filtro_ids = [
        int(i) for i in request.GET.getlist("aluno_filtro") if i.isdigit()
    ]

    aulas_queryset = Aula.objects.select_related("modalidade").prefetch_related(
        "alunos", "professores", "aulas_validadas_por_mim", "presencas_registradas"
    )

    if data_inicial_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            aulas_queryset = aulas_queryset.filter(data_hora__date__gte=data_inicial)
        except ValueError:
            pass
    if data_final_str:
        try:
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
            aulas_queryset = aulas_queryset.filter(data_hora__date__lte=data_final)
        except ValueError:
            pass
    if professor_filtro_id:
        aulas_queryset = aulas_queryset.filter(
            Q(professores__id=professor_filtro_id)
            | Q(relatorioaula__professor_que_validou__id=professor_filtro_id)
        ).distinct()
    if modalidade_filtro_id:
        aulas_queryset = aulas_queryset.filter(modalidade_id=modalidade_filtro_id)
    if aluno_filtro_ids:
        aulas_queryset = aulas_queryset.filter(
            alunos__id__in=aluno_filtro_ids
        ).distinct()

    # --- 2. KPIs ---
    total_aulas = aulas_queryset.count()
    total_realizadas = aulas_queryset.filter(status="Realizada").count()
    total_canceladas = aulas_queryset.filter(status="Cancelada").count()
    total_aluno_ausente = aulas_queryset.filter(status="Aluno Ausente").count()
    aulas_concluidas = total_realizadas + total_aluno_ausente + total_canceladas
    taxa_sucesso = (
        (total_realizadas / aulas_concluidas * 100) if aulas_concluidas > 0 else 0
    )

    # --- 3. Gráficos ---
    # Evolução de aulas por mês
    aulas_por_mes = (
        aulas_queryset.filter(status="Realizada")
        .annotate(mes=TruncMonth("data_hora"))
        .values("mes")
        .annotate(contagem=Count("id"))
        .order_by("mes")
    )
    mes_chart_labels = [item["mes"].strftime("%b/%Y") for item in aulas_por_mes]
    mes_chart_data = [item["contagem"] for item in aulas_por_mes]

    # Volume por categoria
    cat_chart_data_qs = (
        aulas_queryset.filter(modalidade__isnull=False)
        .values("modalidade__nome")
        .annotate(contagem=Count("id"))
        .order_by("-contagem")
    )
    cat_chart_labels = [item["modalidade__nome"].title() for item in cat_chart_data_qs]
    cat_chart_data = [item["contagem"] for item in cat_chart_data_qs]

    # Desempenho modalidades por mês
    dados_agrupados = (
        aulas_queryset.filter(status="Realizada")
        .annotate(mes=TruncMonth("data_hora"))
        .values("mes", "modalidade__nome")
        .annotate(contagem=Count("id"))
        .order_by("mes")
    )

    meses = sorted(set(d["mes"].strftime("%b/%Y") for d in dados_agrupados))
    modalidades = sorted(set(d["modalidade__nome"] for d in dados_agrupados))
    dados_pivotados = {mod: {mes: 0 for mes in meses} for mod in modalidades}
    for item in dados_agrupados:
        mes_str = item["mes"].strftime("%b/%Y")
        dados_pivotados[item["modalidade__nome"]][mes_str] = item["contagem"]

    cores_grafico = [
        "#4e73df",
        "#1cc88a",
        "#36b9cc",
        "#f6c23e",
        "#e74a3b",
        "#858796",
        "#5a5c69",
        "#fd7e14",
    ]
    desempenho_modalidades_datasets = [
        {
            "label": mod,
            "data": list(dados_pivotados[mod].values()),
            "backgroundColor": cores_grafico[i % len(cores_grafico)],
        }
        for i, mod in enumerate(modalidades)
    ]

    # --- 4. Professores ---
    professores_queryset = CustomUser.objects.filter(
        tipo__in=["professor", "admin"]
    ).annotate(
        total_atribuidas=Count(
            "aulas_lecionadas",
            distinct=True,
            filter=Q(aulas_lecionadas__in=aulas_queryset),
        ),
        realizadas_normal=Count(
            "aulas_validadas_por_mim",
            distinct=True,
            filter=Q(
                aulas_validadas_por_mim__aula__in=aulas_queryset,
                aulas_validadas_por_mim__aula__status="Realizada",
            )
            & ~Q(
                aulas_validadas_por_mim__aula__modalidade__nome__icontains="atividade complementar"
            ),
        ),
        realizadas_ac=Count(
            "presencas_registradas",
            distinct=True,
            filter=Q(
                presencas_registradas__aula__in=aulas_queryset,
                presencas_registradas__status="presente",
                presencas_registradas__aula__modalidade__nome__icontains="atividade complementar",
            ),
        ),
    )

    professores_tabela = []
    for p in professores_queryset:
        p.total_realizadas = p.realizadas_normal + p.realizadas_ac
        p.taxa_realizacao = (
            (p.total_realizadas / p.total_atribuidas * 100)
            if p.total_atribuidas > 0
            else 0
        )
        if p.total_atribuidas > 0 or p.total_realizadas > 0:
            professores_tabela.append(p)
    professores_tabela.sort(key=lambda x: x.total_realizadas, reverse=True)

    # --- 5. Contexto ---
    contexto = {
        "titulo": "Relatórios de Aulas",
        "data_inicial": data_inicial_str,
        "data_final": data_final_str,
        "professor_filtro": professor_filtro_id,
        "modalidade_filtro": modalidade_filtro_id,
        "status_filtro": status_filtro,
        "aluno_filtro_ids": aluno_filtro_ids,
        "professores_list": CustomUser.objects.filter(
            tipo__in=["professor", "admin"]
        ).order_by("username"),
        "modalidades_list": Modalidade.objects.all().order_by("nome"),
        "alunos_list": Aluno.objects.all().order_by("nome_completo"),
        "status_choices": Aula.STATUS_AULA_CHOICES,
        "total_aulas": total_aulas,
        "total_realizadas": total_realizadas,
        "total_canceladas": total_canceladas,
        "total_aluno_ausente": total_aluno_ausente,
        "taxa_sucesso": f"{taxa_sucesso:.1f}",
        "cat_chart_labels": cat_chart_labels,
        "cat_chart_data": cat_chart_data,
        "mes_chart_labels": mes_chart_labels,
        "mes_chart_data": mes_chart_data,
        "aulas_por_professor": professores_tabela,
        "desempenho_modalidades_labels": meses,
        "desempenho_modalidades_datasets": desempenho_modalidades_datasets,
    }

    return render(request, "scheduler/relatorios_aulas.html", contexto)


# --- NOVA VIEW DE EXPORTAÇÃO ---
@user_passes_test(is_admin)
def exportar_relatorio_agregado(request):
    """
    Exporta um relatório gerencial AGREGADO e ESTILIZADO para um arquivo Excel (.xlsx),
    respeitando os filtros aplicados na página de relatórios.
    """
    # 1. Lógica de filtros (AGORA CORRIGIDA E COMPLETA)
    aulas_queryset = Aula.objects.all()

    # --- INÍCIO DO BLOCO DE FILTROS CORRIGIDO ---
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get("professor_filtro", "")
    modalidade_filtro_id = request.GET.get("modalidade_filtro", "")
    status_filtro = request.GET.get("status_filtro", "")
    aluno_filtro_ids = request.GET.getlist("aluno_filtro")

    if data_inicial_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__gte=datetime.strptime(
                    data_inicial_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__lte=datetime.strptime(
                    data_final_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if professor_filtro_id:
        aulas_queryset = aulas_queryset.filter(
            Q(professores__id=professor_filtro_id)
            | Q(relatorioaula__professor_que_validou__id=professor_filtro_id)
        )
    if modalidade_filtro_id:
        aulas_queryset = aulas_queryset.filter(modalidade_id=modalidade_filtro_id)
    if status_filtro:
        if status_filtro == "Substituído":
            aulas_queryset = aulas_queryset.filter(
                status="Realizada",
                professores__isnull=False,
                relatorioaula__professor_que_validou__isnull=False,
            ).exclude(professores=F("relatorioaula__professor_que_validou"))
        else:
            aulas_queryset = aulas_queryset.filter(status=status_filtro)
    if aluno_filtro_ids:
        aluno_filtro_ids = [int(id) for id in aluno_filtro_ids if id.isdigit()]
        if aluno_filtro_ids:
            aulas_queryset = aulas_queryset.filter(
                alunos__id__in=aluno_filtro_ids
            ).distinct()
    # --- FIM DO BLOCO DE FILTROS CORRIGIDO ---

    # 2. Cálculo dos KPIs Gerais
    total_aulas = aulas_queryset.count()
    total_realizadas = aulas_queryset.filter(status="Realizada").count()
    total_canceladas = aulas_queryset.filter(status="Cancelada").count()
    total_aluno_ausente = aulas_queryset.filter(status="Aluno Ausente").count()

    # 3. Cálculo dos dados por Professor (Agora sobre o queryset filtrado)
    professores = (
        CustomUser.objects.filter(tipo__in=["professor", "admin"])
        .annotate(
            total_atribuidas=Count(
                "aulas_lecionadas",
                distinct=True,
                filter=Q(aulas_lecionadas__in=aulas_queryset),
            ),
            total_realizadas=Count(
                "aulas_validadas_por_mim",
                distinct=True,
                filter=Q(
                    aulas_validadas_por_mim__aula__in=aulas_queryset,
                    aulas_validadas_por_mim__aula__status="Realizada",
                ),
            ),
            total_ausencias=Count(
                "aulas_validadas_por_mim",
                distinct=True,
                filter=Q(
                    aulas_validadas_por_mim__aula__in=aulas_queryset,
                    aulas_validadas_por_mim__aula__status="Aluno Ausente",
                ),
            ),
        )
        .filter(Q(total_atribuidas__gt=0) | Q(total_realizadas__gt=0))
        .order_by("-total_realizadas")
    )

    # 4. Cálculo dos dados por Categoria (Agora sobre o queryset filtrado)
    aulas_por_modalidade = (
        aulas_queryset.filter(modalidade__isnull=False)
        .values("modalidade__id", "modalidade__nome")
        .annotate(
            total_aulas=Count("id"),
            aulas_realizadas=Count("id", filter=Q(status="Realizada")),
            aulas_ausencias=Count("id", filter=Q(status="Aluno Ausente")),
            aulas_canceladas=Count("id", filter=Q(status="Cancelada")),
        )
        .order_by("-total_aulas")
    )

    # 5. Criação e Estilização da Planilha Excel
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Relatorio Gerencial"

    # Estilos
    font_bold = Font(bold=True)
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(
        start_color="2F75B5", end_color="2F75B5", fill_type="solid"
    )
    fill_subheader = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )

    # --- Bloco de KPIs Gerais ---
    ws.merge_cells("B2:E2")
    cell_kpi_header = ws["B2"]
    cell_kpi_header.value = "Resumo Geral do Período"
    cell_kpi_header.font = font_header
    cell_kpi_header.fill = fill_header
    cell_kpi_header.alignment = Alignment(horizontal="center")

    ws["B4"] = "Total de Aulas no Período:"
    ws["B4"].font = font_bold
    ws["C4"] = total_aulas
    ws["B5"] = "Aulas Realizadas:"
    ws["B5"].font = font_bold
    ws["C5"] = total_realizadas
    ws["B6"] = "Aulas com Ausência:"
    ws["B6"].font = font_bold
    ws["C6"] = total_aluno_ausente
    ws["B7"] = "Aulas Canceladas:"
    ws["B7"].font = font_bold
    ws["C7"] = total_canceladas

    # --- Tabela de Resumo por Professor ---
    current_row = 10
    ws.merge_cells(f"B{current_row}:F{current_row}")
    cell_prof_header = ws[f"B{current_row}"]
    cell_prof_header.value = "Resumo por Professor"
    cell_prof_header.font = font_header
    cell_prof_header.fill = fill_header
    cell_prof_header.alignment = Alignment(horizontal="center")
    current_row += 1

    prof_headers = [
        "Professor",
        "Aulas Atribuídas",
        "Aulas Realizadas",
        "Aulas c/ Ausência",
        "Taxa de Realização (%)",
    ]
    ws.append(
        [""] + prof_headers
    )  # Adiciona uma coluna vazia no início para espaçamento
    for cell in ws[current_row]:
        cell.font = font_bold
        cell.fill = fill_subheader

    current_row += 1
    for p in professores:
        taxa = (
            (p.total_realizadas / p.total_atribuidas * 100)
            if p.total_atribuidas > 0
            else 0
        )
        ws.append(
            [
                "",
                p.username.title(),
                p.total_atribuidas,
                p.total_realizadas,
                p.total_ausencias,
                f"{taxa:.2f}%",
            ]
        )

    # --- Tabela de Resumo por Categoria ---
    current_row += 3  # Espaço entre as tabelas
    ws.merge_cells(f"B{current_row}:G{current_row}")
    cell_cat_header = ws[f"B{current_row}"]
    cell_cat_header.value = "Resumo por Categoria"
    cell_cat_header.font = font_header
    cell_cat_header.fill = fill_header
    cell_cat_header.alignment = Alignment(horizontal="center")
    current_row += 1

    cat_headers = [
        "Categoria",
        "Total de Aulas",
        "Aulas Realizadas",
        "Ausências",
        "Canceladas",
        "Taxa de Realização (%)",
    ]
    ws.append([""] + cat_headers)
    for cell in ws[current_row]:
        cell.font = font_bold
        cell.fill = fill_subheader

    current_row += 1
    for item in aulas_por_modalidade:
        taxa = (
            (item["aulas_realizadas"] / item["total_aulas"] * 100)
            if item["total_aulas"] > 0
            else 0
        )
        ws.append(
            [
                "",
                item["modalidade__nome"].title(),
                item["total_aulas"],
                item["aulas_realizadas"],
                item["aulas_ausencias"],
                item["aulas_canceladas"],
                f"{taxa:.2f}%",
            ]
        )

    # Ajuste final das colunas
    for i, column_cells in enumerate(ws.columns):
        max_length = 0
        column = get_column_letter(i + 1)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # 6. Preparação da resposta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Resumo Gerencial.xlsx"'},
    )
    workbook.save(response)

    return response


# --- NOVA VIEW PARA EXPORTAÇÃO DE DADOS ---
@user_passes_test(is_admin)
def exportar_aulas(request):
    """
    Exporta um relatório detalhado de atividades das aulas para um arquivo CSV,
    respeitando os filtros aplicados. Cada linha representa uma atividade.
    """
    # 1. Reaplica a lógica de filtros da página de listagem (nenhuma mudança aqui)
    aulas_queryset = Aula.objects.all()
    # (Toda a sua lógica de filtros por data, professor, aluno, etc., permanece a mesma)
    data_inicial_str = request.GET.get("data_inicial", "")
    data_final_str = request.GET.get("data_final", "")
    professor_filtro_id = request.GET.get("professor_filtro", "")
    modalidade_filtro_id = request.GET.get("modalidade_filtro", "")
    status_filtro = request.GET.get("status_filtro", "")
    aluno_filtro_ids = request.GET.getlist("aluno_filtro")

    if data_inicial_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__gte=datetime.strptime(
                    data_inicial_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if data_final_str:
        try:
            aulas_queryset = aulas_queryset.filter(
                data_hora__date__lte=datetime.strptime(
                    data_final_str, "%Y-%m-%d"
                ).date()
            )
        except ValueError:
            pass
    if professor_filtro_id:
        aulas_queryset = aulas_queryset.filter(
            Q(professores__id=professor_filtro_id)
            | Q(relatorioaula__professor_que_validou__id=professor_filtro_id)
        )
    if modalidade_filtro_id:
        aulas_queryset = aulas_queryset.filter(modalidade_id=modalidade_filtro_id)
    if status_filtro:
        if status_filtro == "Substituído":
            aulas_queryset = aulas_queryset.filter(
                status="Realizada",
                professores__isnull=False,
                relatorioaula__professor_que_validou__isnull=False,
            ).exclude(professores=F("relatorioaula__professor_que_validou"))
        else:
            aulas_queryset = aulas_queryset.filter(status=status_filtro)
    if aluno_filtro_ids:
        aluno_filtro_ids = [int(id) for id in aluno_filtro_ids if id.isdigit()]
        if aluno_filtro_ids:
            aulas_queryset = aulas_queryset.filter(
                alunos__id__in=aluno_filtro_ids
            ).distinct()

    # 2. Otimiza a consulta para performance, buscando todos os dados relacionados de uma vez
    aulas_list = list(
        aulas_queryset.order_by("data_hora")
        .select_related("modalidade", "relatorioaula__professor_que_validou")
        .prefetch_related(
            "alunos",
            "professores",
            "relatorioaula__itens_rudimentos",
            "relatorioaula__itens_ritmo",
            "relatorioaula__itens_viradas",
        )
    )

    # 3. Workbook e Estilos...
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Relatorio Detalhado de Aulas"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    cores_atividades = {
        "Teoria": PatternFill(
            start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        ),
        "Repertório": PatternFill(
            start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"
        ),
        "Rudimento": PatternFill(
            start_color="EAF1DD", end_color="EAF1DD", fill_type="solid"
        ),
        "Ritmo": PatternFill(
            start_color="DBEEF3", end_color="DBEEF3", fill_type="solid"
        ),
        "Virada": PatternFill(
            start_color="E5E0EC", end_color="E5E0EC", fill_type="solid"
        ),
        "Observações Gerais": PatternFill(
            start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"
        ),
        "N/A": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    # 4. Cabeçalho
    headers = [
        "ID Aula",
        "Data e Hora",
        "Status",
        "Alunos",
        "Prof. Atribuído(s)",
        "Prof. Realizou",
        "Categoria",
        "Tipo de Conteúdo",
        "Descrição",
        "Detalhes (BPM, Livro, Duração)",
        "Observações do Conteúdo",
    ]
    worksheet.append(headers)
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # 5. Loop principal com a lógica corrigida
    for i, aula in enumerate(aulas_list):
        if i > 0:
            worksheet.append([])

        alunos_str = ", ".join([al.nome_completo.title() for al in aula.alunos.all()])
        professores_str = ", ".join(
            [p.username.title() for p in aula.professores.all()]
        )
        relatorio = getattr(aula, "relatorioaula", None)
        professor_realizou_str = (
            relatorio.professor_que_validou.username.title()
            if relatorio and relatorio.professor_que_validou
            else "N/A"
        )
        data_hora_naive = timezone.localtime(aula.data_hora).replace(tzinfo=None)
        base_row_data = [
            aula.id,
            data_hora_naive,
            aula.get_status_display(),
            alunos_str,
            professores_str,
            professor_realizou_str,
            aula.modalidade.nome,
        ]

        def adicionar_linha_estilizada(dados_atividade):
            worksheet.append(base_row_data + dados_atividade)
            cor_tipo = dados_atividade[0]
            fill_style = cores_atividades.get(cor_tipo, cores_atividades["N/A"])
            for cell in worksheet[worksheet.max_row]:
                cell.fill = fill_style

        if relatorio:
            conteudo_adicionado = False
            if relatorio.conteudo_teorico:
                adicionar_linha_estilizada(
                    [
                        "Teoria",
                        relatorio.conteudo_teorico,
                        "",
                        relatorio.observacoes_teoria or "",
                    ]
                )
                conteudo_adicionado = True
            for item in relatorio.itens_rudimentos.all():
                detalhes = f"BPM: {item.bpm or 'N/A'} / Duração: {item.duracao_min or 'N/A'} min"
                adicionar_linha_estilizada(
                    ["Rudimento", item.descricao, detalhes, item.observacoes or ""]
                )
                conteudo_adicionado = True
            for item in relatorio.itens_ritmo.all():
                detalhes = f"Livro: {item.livro_metodo or 'N/A'} / BPM: {item.bpm or 'N/A'} / Duração: {item.duracao_min or 'N/A'} min"
                adicionar_linha_estilizada(
                    ["Ritmo", item.descricao, detalhes, item.observacoes or ""]
                )
                conteudo_adicionado = True
            for item in relatorio.itens_viradas.all():
                detalhes = f"BPM: {item.bpm or 'N/A'} / Duração: {item.duracao_min or 'N/A'} min"
                adicionar_linha_estilizada(
                    ["Virada", item.descricao, detalhes, item.observacoes or ""]
                )
                conteudo_adicionado = True
            if relatorio.repertorio_musicas:
                adicionar_linha_estilizada(
                    [
                        "Repertório",
                        relatorio.repertorio_musicas,
                        "",
                        relatorio.observacoes_repertorio or "",
                    ]
                )
                conteudo_adicionado = True
            if relatorio.observacoes_gerais:
                adicionar_linha_estilizada(
                    ["Observações Gerais", relatorio.observacoes_gerais, "", ""]
                )
                conteudo_adicionado = True
            if not conteudo_adicionado:
                adicionar_linha_estilizada(
                    ["N/A", "Relatório existe, mas está vazio.", "", ""]
                )
        else:
            adicionar_linha_estilizada(["N/A", "Aula sem relatório criado.", "", ""])

    # 7. Ajuste de colunas
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20
    worksheet.column_dimensions["D"].width = 40
    worksheet.column_dimensions["I"].width = 40
    worksheet.column_dimensions["J"].width = 50
    worksheet.column_dimensions["K"].width = 50

    # 8. Retorno da resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="Relatório de Aulas.xlsx"'
        },
    )
    workbook.save(response)
    return response


@login_required
def get_horarios_ocupados(request):
    professor_id = request.GET.get("professor_id")
    data_selecionada_str = request.GET.get(
        "data"
    )  # Data selecionada no calendário (YYYY-MM-DD)
    aula_id = request.GET.get("aula_id")  # ID da aula atual sendo editada (se houver)

    horarios_ocupados = (
        []
    )  # Lista de strings de horários (HH:MM) para o dia selecionado
    dias_com_aulas = []  # Lista de strings de datas (YYYY-MM-DD) para o professor

    if professor_id:  # A busca por dias com aulas não precisa de uma data específica
        try:
            professor_id = int(professor_id)
            aula_id = int(aula_id) if aula_id else None

            # --- Buscar todos os dias com aulas para o professor selecionado ---
            # Filtra aulas para o professor
            aulas_do_professor = (
                Aula.objects.filter(professor_id=professor_id)
                .values_list("data_hora__date", flat=True)
                .distinct()
            )  # Pega apenas as datas únicas

            if aula_id:  # Exclui a própria aula se estiver editando
                aulas_do_professor = aulas_do_professor.exclude(id=aula_id)

            for data_aula in aulas_do_professor:
                dias_com_aulas.append(data_aula.strftime("%Y-%m-%d"))

            # --- Buscar horários ocupados para o dia específico (se fornecido) ---
            if data_selecionada_str:
                data_selecionada = datetime.strptime(
                    data_selecionada_str, "%Y-%m-%d"
                ).date()

                aulas_do_dia = Aula.objects.filter(
                    professor_id=professor_id, data_hora__date=data_selecionada
                )
                if aula_id:  # Exclui a própria aula se estiver editando
                    aulas_do_dia = aulas_do_dia.exclude(id=aula_id)

                for aula in aulas_do_dia:
                    horarios_ocupados.append(localtime(aula.data_hora).strftime("%H:%M"))

            return JsonResponse(
                {
                    "horarios_ocupados": horarios_ocupados,
                    "dias_com_aulas": dias_com_aulas,
                }
            )

        except (ValueError, TypeError) as e:
            return JsonResponse(
                {
                    "horarios_ocupados": [],
                    "dias_com_aulas": [],
                    "erro": f"Dados inválidos: {e}",
                },
                status=400,
            )

    return JsonResponse(
        {
            "horarios_ocupados": [],
            "dias_com_aulas": [],
            "erro": "Professor não fornecido.",
        },
        status=400,
    )


#  --- NOVA VIEW PARA OBTER EVENTOS DO CALENDÁRIO (FULLCALENDAR) ---
@login_required
def get_eventos_calendario(request):
    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    professor_filtro_id = request.GET.get("professor_filtro_id", "")

    events = []

    if not (start_str and end_str):
        return JsonResponse({"error": "Período não fornecido."}, status=400)

    try:
        start_date = datetime.fromisoformat(start_str.replace("Z", "+00:00")).date()
        end_date = datetime.fromisoformat(end_str.replace("Z", "+00:00")).date()

        aulas_no_periodo = Aula.objects.filter(
            data_hora__date__range=(start_date, end_date)
        )

        # Filtro por professor
        if request.user.tipo == "professor":
            aulas_no_periodo = aulas_no_periodo.filter(professores=request.user)
        elif request.user.tipo in ["admin", "comercial"] and professor_filtro_id:
            aulas_no_periodo = aulas_no_periodo.filter(
                professores__id=professor_filtro_id
            )

        # --- OTIMIZAÇÃO E CORREÇÃO PRINCIPAL ---
        # Usamos prefetch_related para buscar todos os alunos e professores de uma vez,
        # evitando múltiplas queries ao banco de dados e melhorando a performance.
        aulas_no_periodo = (
            aulas_no_periodo.select_related(
                "modalidade", "relatorioaula", "relatorioaula__professor_que_validou"
            )
            .prefetch_related("alunos", "professores")
            .distinct()
        )

        for aula in aulas_no_periodo:
            # --- NOVA LÓGICA PARA LIDAR COM MÚLTIPLOS ALUNOS/PROFESSORES ---
            alunos_list = list(aula.alunos.all())
            professores_list = list(aula.professores.all())

            title = ""
            aluno_prop_str = ""

            # Define o título do evento e a string de alunos para o popover
            if len(alunos_list) == 1:
                # Se há apenas um aluno, o título é o nome dele
                title = f"{alunos_list[0].nome_completo.split()[0].title()} ({aula.modalidade.nome})"
                aluno_prop_str = alunos_list[0].nome_completo
            elif len(alunos_list) > 1:
                # Se for em grupo, o título é genérico
                title = f"Grupo: {aula.modalidade.nome}"
                aluno_prop_str = f"{len(alunos_list)} alunos"
            else:
                # Se não houver alunos (Ex: Atividade Complementar)
                title = aula.modalidade.nome
                aluno_prop_str = "Nenhum aluno"

            # Junta os nomes dos professores para o popover
            prof_atribuido_str = (
                ", ".join([p.username.title() for p in professores_list]) or "N/A"
            )

            event_class = f'status-{aula.status.replace(" ", "")}'
            professor_realizou = getattr(aula, "relatorioaula", None) and getattr(
                aula.relatorioaula, "professor_que_validou", None
            )

            events.append(
                {
                    "title": title,
                    "start": aula.data_hora.isoformat(),
                    "url": f"/aula/{aula.pk}/validar/",  # A URL pode ser a de validar/ver relatório
                    "classNames": [event_class],
                    "extendedProps": {
                        "status": aula.status,
                        "aluno": aluno_prop_str,
                        "professor_atribuido": prof_atribuido_str,
                        "professor_realizou": (
                            professor_realizou.username.title()
                            if professor_realizou
                            else "N/A"
                        ),
                        "modalidade": aula.modalidade.nome.title(),
                    },
                }
            )

        return JsonResponse(events, safe=False)

    except (ValueError, TypeError) as e:
        # Em caso de erro, é útil registrar o erro no console do servidor para depuração
        print(f"Erro em get_eventos_calendario: {e}")
        return JsonResponse(
            {"error": "Erro interno ao processar a requisição."}, status=500
        )


@login_required
def perfil_usuario(request):
    user = request.user
    if request.method == "POST":
        form = UserProfileForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Seu perfil foi atualizado com sucesso!")
            return redirect(
                "scheduler:perfil_usuario"
            )  # Redireciona de volta para a página de perfil
        else:
            messages.error(request, "Erro ao atualizar seu perfil. Verifique os dados.")
    else:
        form = UserProfileForm(
            instance=user
        )  # Preenche o formulário com os dados do usuário

    contexto = {"form": form, "user_obj": user, "titulo": "Meu Perfil"}
    return render(request, "scheduler/perfil_usuario.html", contexto)


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "comercial"])
def get_horario_fixo_data(request):
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    today = timezone.now().date()
    start_date = today - timedelta(weeks=6)

    aulas_periodo = Aula.objects.filter(
        data_hora__date__gte=start_date,
        data_hora__date__lte=today,
        status__in=["Realizada", "Agendada"],
    ).prefetch_related("alunos", "professores").select_related("modalidade")

    ocorrencias_aluno = defaultdict(list)
    for aula in aulas_periodo:
        if not aula.alunos.exists():
            continue
        dt_local = localtime(aula.data_hora)
        dia_semana = dt_local.weekday()
        horario = dt_local.strftime("%H:00")
        data_aula = dt_local.date()

        # Professores formatados igual aos alunos
        conectivos = {"da", "de", "do", "das", "dos", "di", "du"}
        def formatar_nome(nome):
            partes = nome.strip().split()
            if not partes:
                return "?"
            elif len(partes) == 1:
                return partes[0].capitalize()
            elif partes[1].lower() in conectivos and len(partes) > 2:
                p1 = partes[0].capitalize()
                p2 = partes[1].lower()
                p3 = partes[2].capitalize()
                return f"{p1} {p2} {p3}"
            else:
                return f"{partes[0].capitalize()} {partes[1].capitalize()}"

        prof_nomes = [formatar_nome(p.username) for p in aula.professores.all()]
        modalidade_nome = aula.modalidade.nome.title()

        payload = {
            "slot": (dia_semana, horario),
            "data": data_aula,
            "profs": prof_nomes,
            "modalidade": modalidade_nome
        }
        for aluno in aula.alunos.all():
            ocorrencias_aluno[aluno.id].append(payload)

    horario_principal_aluno = {}
    for aluno_id, aulas in ocorrencias_aluno.items():
        frequencia_slots = defaultdict(int)
        data_recente_slot = {}
        payload_recente_slot = {}
        for aula_data in aulas:
            slot = aula_data["slot"]
            frequencia_slots[slot] += 1
            if aula_data["data"] > data_recente_slot.get(slot, date.min):
                data_recente_slot[slot] = aula_data["data"]
                payload_recente_slot[slot] = aula_data
        if frequencia_slots:
            slot_principal = max(
                frequencia_slots.keys(),
                key=lambda slot: (frequencia_slots[slot], data_recente_slot[slot]),
            )
            payload_principal = payload_recente_slot[slot_principal]
            horario_principal_aluno[aluno_id] = {
                "slot": slot_principal,
                "contagem": frequencia_slots[slot_principal],
                "profs": payload_principal["profs"],
                "modalidade": payload_principal["modalidade"]
            }

    aluno_ids = list(horario_principal_aluno.keys())
    alunos_map = {}
    conectivos = {"da", "de", "do", "das", "dos", "di", "du"}

    def formatar_nome(nome):
        partes = nome.strip().split()
        if not partes:
            return "Aluno ?"
        elif len(partes) == 1:
            return partes[0].capitalize()
        elif partes[1].lower() in conectivos and len(partes) > 2:
            p1 = partes[0].capitalize()
            p2 = partes[1].lower()
            p3 = partes[2].capitalize()
            return f"{p1} {p2} {p3}"
        else:
            return f"{partes[0].capitalize()} {partes[1].capitalize()}"

    for aluno in Aluno.objects.filter(id__in=aluno_ids):
        alunos_map[aluno.id] = formatar_nome(aluno.nome_completo)

    LIMITE_HORARIO_FIXO = 3
    grade_horarios = defaultdict(dict)

    for aluno_id, info in horario_principal_aluno.items():
        dia_semana, horario = info["slot"]
        contagem = info["contagem"]
        aluno_nome = alunos_map.get(aluno_id, "Aluno ?")
        status_aluno = "fixo" if contagem >= LIMITE_HORARIO_FIXO else "variavel"
        prof_nomes = info["profs"]
        modalidade = info["modalidade"]
        if not grade_horarios[horario].get(dia_semana):
            grade_horarios[horario][dia_semana] = {
                "status": status_aluno,
                "alunos": {aluno_nome},
                "profs": set(prof_nomes),
                "modalidades": {modalidade}
            }
        else:
            grade_horarios[horario][dia_semana]["alunos"].add(aluno_nome)
            grade_horarios[horario][dia_semana]["profs"].update(prof_nomes)
            grade_horarios[horario][dia_semana]["modalidades"].add(modalidade)
            if status_aluno == "fixo":
                grade_horarios[horario][dia_semana]["status"] = "fixo"

    for horario, dias in grade_horarios.items():
        for dia, slot in dias.items():
            slot["alunos_texto"] = "\n".join(sorted(slot["alunos"]))
            slot["profs_texto"] = "\n".join(sorted(slot["profs"]))
            slot["modalidades_texto"] = "\n".join(sorted(slot["modalidades"]))
            del slot["alunos"]
            del slot["profs"]
            del slot["modalidades"]

    if is_ajax:
        return JsonResponse(grade_horarios)
    else:
        total_fixo = 0
        total_variavel = 0
        aulas_por_dia = [0] * 6
        horarios_visiveis = [f"{h:02d}:00" for h in range(8, 21)]
        horarios_intervalo = ['12:00', '13:00']
        slots_intervalo_livres = 0

        for horario in horarios_visiveis:
            for dia_index in range(6):
                slot_info = grade_horarios[horario].get(dia_index)
                if slot_info:
                    if slot_info["status"] == "fixo":
                        total_fixo += 1
                    else:
                        total_variavel += 1
                    aulas_por_dia[dia_index] += 1
                else:
                    if horario in horarios_intervalo:
                        slots_intervalo_livres += 1

        total_slots_disponiveis = 6 * 13
        total_ocupados = total_fixo + total_variavel
        total_livres = (total_slots_disponiveis - total_ocupados) - slots_intervalo_livres
        total_slots_agendaveis = total_slots_disponiveis - slots_intervalo_livres
        taxa_ocupacao = (
            (total_ocupados / total_slots_agendaveis) * 100
            if total_slots_agendaveis > 0
            else 0
        )

        dias_semana_nomes = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
        ja_viu_tour = request.user.tours_vistos.filter(tour_id="horarios_fixos_v1").exists()

        context = {
            "titulo": "Grade de Horários",
            "mostrar_tour_horarios": not ja_viu_tour,
            "total_fixo": total_fixo,
            "total_variavel": total_variavel,
            "total_livres": total_livres,
            "taxa_ocupacao": f"{taxa_ocupacao:.1f}",
            "ocupacao_chart_labels": dias_semana_nomes,
            "ocupacao_chart_data": aulas_por_dia,
        }
        return render(request, "scheduler/horarios_grid.html", context)
  

@login_required
@require_POST
def marcar_tour_visto(request):
    """Marca que o usuário logado completou um tour específico."""
    try:
        data = json.loads(request.body)
        tour_id = data.get("tour_id")

        if not tour_id:
            return JsonResponse(
                {"success": False, "error": "tour_id não fornecido"}, status=400
            )

        TourVisto.objects.get_or_create(usuario=request.user, tour_id=tour_id)

        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required
@user_passes_test(lambda u: u.tipo in ["admin", "professor", "comercial"])
def listar_reposicoes_pendentes(request):
    """
    Lista todas as faltas justificadas que ainda não tiveram uma aula de
    reposição agendada.
    """
    reposicoes_pendentes_qs = (
        PresencaAluno.objects.filter(
            status="ausente", tipo_falta="justificada", aula_reposicao__isnull=True
        )
        .select_related("aluno", "aula__modalidade")
        .prefetch_related("aula__professores")
        .order_by("aula__data_hora")
    )

    if request.user.tipo == "professor":
        alunos_do_professor = Aula.objects.filter(professores=request.user).values_list(
            "alunos", flat=True
        )
        reposicoes_pendentes_qs = reposicoes_pendentes_qs.filter(
            aluno_id__in=set(alunos_do_professor)
        )

    contexto = {
        "titulo": "Controle de Reposições Pendentes",
        "reposicoes": reposicoes_pendentes_qs,
    }
    return render(request, "scheduler/reposicoes_listar.html", contexto)


@login_required
@require_POST
def gerar_relatorio_anual_ia(request, aluno_id):
    try:
        aluno = Aluno.objects.get(pk=aluno_id)
        ano_atual = 2025

        aulas = Aula.objects.filter(
            alunos=aluno,
            data_hora__year=ano_atual,
            status='Realizada'
        ).exclude(relatorioaula__isnull=True).order_by('data_hora')

        if not aulas.exists():
            return JsonResponse({
                'status': 'error',
                'message': f'Nenhum relatório encontrado para {aluno.nome_completo} em {ano_atual}.'
            })

        historico_aulas = []
        rudimentos_stats = defaultdict(list)
        repertorio_set = set()
        cursos_reais = set()

        for aula in aulas:
            if aula.modalidade:
                cursos_reais.add(aula.modalidade.nome)

            rel = getattr(aula, 'relatorioaula', None)
            if not rel:
                continue

            rudimentos_txt_aula = []
            for item in rel.itens_rudimentos.all():
                rudimentos_txt_aula.append(f"{item.descricao} ({item.bpm}bpm)")
                if item.bpm:
                    nums = re.findall(r'\d+', str(item.bpm))
                    if nums:
                        val = int(max(nums, key=int))
                        rudimentos_stats[item.descricao.strip().title()].append(val)

            if rel.repertorio_musicas:
                musicas = re.split(r'[,\n]+', rel.repertorio_musicas)
                for m in musicas:
                    m_limpa = m.strip()
                    if m_limpa and len(m_limpa) > 2:
                        repertorio_set.add(m_limpa)

            dados_aula = []
            if rel.conteudo_teorico:
                dados_aula.append(f"Teoria: {rel.conteudo_teorico}")
            if rudimentos_txt_aula:
                dados_aula.append(f"Rudimentos: {', '.join(rudimentos_txt_aula)}")
            if rel.repertorio_musicas:
                dados_aula.append(f"Repertório: {rel.repertorio_musicas}")
            if rel.observacoes_gerais:
                dados_aula.append(f"Obs: {rel.observacoes_gerais}")

            if dados_aula:
                data_str = aula.data_hora.strftime('%d/%m')
                nome_modalidade = aula.modalidade.nome if aula.modalidade else "Aula"
                conteudo_str = " | ".join(dados_aula)
                historico_aulas.append(f"### {data_str} ({nome_modalidade}):\n   {conteudo_str}")

        texto_historico = "\n\n".join(historico_aulas)

        curso_str = ", ".join(cursos_reais) if cursos_reais else "Curso não definido"

        lista_evolucao = []
        for nome, valores in rudimentos_stats.items():
            if valores:
                mini = min(valores)
                maxi = max(valores)
                delta = maxi - mini 
                
                if delta > 0: 
                    lista_evolucao.append({'nome': nome, 'min': mini, 'max': maxi, 'delta': delta})

        top_5_rudimentos = sorted(lista_evolucao, key=lambda x: x['delta'], reverse=True)[:5]

        lista_rudimentos_stats = ""
        for r in top_5_rudimentos:
            lista_rudimentos_stats += f"- {r['nome']}: Início {r['min']}bpm -> Máx {r['max']}bpm (Ganho de +{r['delta']}bpm)\n"

        if not lista_rudimentos_stats:
            lista_rudimentos_stats = "Nenhum dado numérico de BPM suficiente para cálculo de evolução."

        lista_repertorio = "\n".join([f"- {m}" for m in repertorio_set]) or "Nenhuma música específica registrada."

        prompt = f"""
        Atue como Coordenador Pedagógico do *Studio Batucada*.  
        Escreva **exclusivamente o corpo do relatório anual**, em **formato Markdown**, seguindo rigorosamente todas as regras abaixo.

        ---

        # 🔹 DADOS DO ALUNO
        - Nome: **{aluno.nome_completo}**
        - Curso: **{curso_str}**
        - Ano: **{ano_atual}**

        # 🔹 ESTATÍSTICAS DE EVOLUÇÃO (dados brutos)
        {lista_rudimentos_stats}

        # 🔹 REPERTÓRIO REGISTRADO
        {lista_repertorio}

        # 🔹 HISTÓRICO DE AULAS (contexto)
        {texto_historico}

        ---

        # 🚨 REGRAS OBRIGATÓRIAS

        ## 1) Estrutura fixa (NÃO ALTERAR)
        O relatório **deve seguir exatamente** esta estrutura:

        ## 1. Visão Geral  
        (texto narrativo)

        ## 2. Teoria e Leitura  
        (texto narrativo)

        ## 3. Técnica e Rudimentos  
        (texto narrativo)

        ###Destaques de Evolução (Top 5):
        {lista_rudimentos_stats}  
        (MANTER esta lista exatamente como está — mesma ordem, mesmos hifens, mesmo conteúdo)

        [GRAFICO_EVOLUCAO]

        ## 4. Coordenação e Ritmos  
        (texto narrativo)

        ## 5. Repertório Musical  
        (texto narrativo)

        ###Músicas Trabalhadas:
        {lista_repertorio}  
        (manter exatamente a lista recebida)

        ## 6. Pontos Fortes  
        (texto narrativo, 2–3 pontos)

        ## 7. Pontos de Melhoria  
        (texto narrativo, 2–3 pontos)

        ## 8. Conclusão  
        (texto narrativo)

        ---

        # 🚨 REGRAS DE FORMATAÇÃO (MUITO IMPORTANTES)

        ### 🔒 Cabeçalho / Rodapé proibidos
        - Não incluir **título**, **nome do aluno no topo**, **data**, **assinatura**, **“Atenciosamente”**, **nomes de responsáveis**.

        ### 🔒 Curso
        - Sempre citar o curso exatamente como: **"{curso_str}"**  
        (idêntico, mesmo o plural, espaçamento e acentuação)

        ### 🔒 Listas 
        Para garantir que as listas fiquem corretamente estruturadas no PDF faça listas usando a marcação para listas quando for listar os rudimentos, músicas, pontos fortes e pontos fracos

        ### ⚠️ NÃO reescrever ou resumir os itens das listas de rudimentos e repertório.
        Apenas reproduzir exatamente os itens enviados.

        ---

        # 🧠 INSTRUÇÃO FINAL
        Escreva com tom profissional, claro e avaliativo, mantendo coerência com os dados do aluno e com o histórico enviado.  
        Não invente números, não altere datas e não descreva conteúdos inexistentes.

        """

        model_name = 'gemini-2.5-flash'

        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            return JsonResponse({'status': 'success', 'relatorio': response.text})
        except Exception as e_model:
            return JsonResponse({'status': 'error', 'message': f"Erro na IA ({model_name}): {str(e_model)}"})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f"Erro interno: {str(e)}"})


def link_callback(uri, rel):
    """
    Converte URLs de arquivos estáticos (ex: /static/img/logo.png)
    em caminhos absolutos do sistema de arquivos (ex: C:/Users/.../static/img/logo.png)
    para que o xhtml2pdf consiga carregar as imagens.
    """
    sUrl = settings.STATIC_URL
    sRoot = settings.STATIC_ROOT
    mUrl = settings.MEDIA_URL
    mRoot = settings.MEDIA_ROOT

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))
    else:
        return uri

    if not os.path.isfile(path):
        result = finders.find(uri.replace(sUrl, ""))
        if result:
            if isinstance(result, (list, tuple)):
                path = result[0]
            else:
                path = result

    if not os.path.isfile(path):
        raise Exception(f'media URI must start with {sUrl} or {mUrl}. Path: {path}')

    return path


@login_required
def baixar_relatorio_pdf(request):
    if request.method == 'POST':
        texto_markdown = request.POST.get('texto_relatorio', '')
        nome_aluno_post = request.POST.get('nome_aluno', 'Aluno')

        aluno = Aluno.objects.filter(nome_completo=nome_aluno_post).first()
        ano_atual = 2025

        curso_str = "Curso não identificado"
        if aluno:
            aulas_ano = Aula.objects.filter(
                alunos=aluno,
                data_hora__year=ano_atual,
                status='Realizada'
            ).select_related('modalidade')

            modalidades = set()
            for aula in aulas_ano:
                if aula.modalidade:
                    modalidades.add(aula.modalidade.nome)

            if modalidades:
                curso_str = ", ".join(modalidades)

        grafico_base64 = None
        if aluno:
            itens = ItemRudimento.objects.filter(
                relatorio__aula__alunos=aluno,
                relatorio__aula__data_hora__year=ano_atual
            ).select_related('relatorio__aula').order_by('relatorio__aula__data_hora')

            dados_rudimentos = defaultdict(list)

            for item in itens:
                if item.bpm:
                    bpm_match = re.findall(r'\d+', str(item.bpm))
                    if bpm_match:
                        bpm_valor = int(max(bpm_match, key=int))
                        chave = item.descricao.strip().title()
                        data_aula = item.relatorio.aula.data_hora.strftime('%d/%m')
                        dados_rudimentos[chave].append((data_aula, bpm_valor))

            top_rudimentos = sorted(dados_rudimentos.items(), key=lambda x: len(x[1]), reverse=True)[:5]

            if top_rudimentos:
                plt.figure(figsize=(10, 4))

                tem_dados = False

                # Jitter horizontal (em índice, pois datas são strings)
                jitter_valores = [-0.15, -0.05, 0.05, 0.15, 0.25]  # até 5 rudimentos
                idx_rudi = 0

                for nome, pontos in top_rudimentos:
                    if len(pontos) >= 1:
                        datas, bpms = zip(*pontos)

                        # converte datas "dd/mm" para índice numérico
                        x_base = list(range(len(datas)))

                        # aplica jitter diferente para cada linha
                        jitter = jitter_valores[idx_rudi % len(jitter_valores)]
                        x_jitter = [x + jitter for x in x_base]

                        plt.plot(
                            x_jitter,
                            bpms,
                            marker='o',
                            linewidth=2,
                            label=nome
                        )

                        tem_dados = True
                        idx_rudi += 1

                if tem_dados:
                    # substitui números do eixo por suas datas originais
                    plt.xticks(range(len(datas)), datas)

                    plt.title(f'Evolução Técnica (BPM) - {ano_atual}', fontsize=12, fontweight='bold')
                    plt.xlabel('Aulas', fontsize=9)
                    plt.ylabel('BPM', fontsize=9)
                    plt.legend(title="Exercícios", fontsize='small')
                    plt.grid(True, linestyle='--', alpha=0.5)
                    plt.tight_layout()

                    buffer = io.BytesIO()
                    plt.savefig(buffer, format='png', transparent=True)
                    buffer.seek(0)
                    image_png = buffer.getvalue()
                    buffer.close()
                    plt.close()

                    grafico_base64 = base64.b64encode(image_png).decode('utf-8')

        html_conteudo = markdown.markdown(texto_markdown)

        html_grafico = ""
        if grafico_base64:
            html_grafico = f"""
            <div style="text-align: center; margin: 30px 0; page-break-inside: avoid;">
                <h4 style="color: #333; border-bottom: 2px solid #ffc107; display: inline-block; margin-bottom: 15px;">
                    Gráfico de Evolução Técnica
                </h4><br>
                <img src="data:image/png;base64,{grafico_base64}" style="width: 100%; max-width: 17cm;">
                <p style="font-size: 9px; color: #666; margin-top: 5px;">
                    * Evolução de velocidade (BPM) nos principais rudimentos praticados.
                </p>
            </div>
            """

        if '[GRAFICO_EVOLUCAO]' in html_conteudo and html_grafico:
            html_conteudo = html_conteudo.replace('[GRAFICO_EVOLUCAO]', html_grafico)
        elif html_grafico:
            if '<h2>Técnica' in html_conteudo:
                partes = html_conteudo.split('<h2>Técnica')
                if len(partes) > 1:
                    subpartes = partes[1].split('<h2>')
                    if len(subpartes) > 1:
                        nova_parte = subpartes[0] + html_grafico
                        html_conteudo = partes[0] + '<h2>Técnica' + nova_parte + '<h2>' + '<h2>'.join(subpartes[1:])
                    else:
                        html_conteudo += f"<br>{html_grafico}"
                else:
                    html_conteudo += f"<br><hr>{html_grafico}"
            else:
                html_conteudo += f"<br><hr>{html_grafico}"

        nome_arquivo_imagem = 'scheduler/img/logo_relatorio.png'
        logo_path = finders.find(nome_arquivo_imagem)

        if not logo_path:
            logo_path = os.path.join(settings.BASE_DIR, 'scheduler', 'static', 'scheduler', 'img', 'logo_relatorio.png')

        context = {
            'conteudo': html_conteudo,
            'aluno': nome_aluno_post,
            'curso': curso_str,
            'ano': ano_atual,
            'logo_path': logo_path,
        }

        template_path = 'scheduler/pdf/relatorio_anual_pdf.html'
        template = get_template(template_path)
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Relatorio {nome_aluno_post} - {ano_atual}.pdf"'

        pisa_status = pisa.CreatePDF(
            html,
            dest=response,
            link_callback=link_callback
        )

        if pisa_status.err:
            return HttpResponse(f'Erro ao gerar PDF: {pisa_status.err}')
        return response

    return HttpResponse("Método não permitido")


@login_required
@user_passes_test(is_admin)
def normalizar_rudimentos(request):
    if request.method == 'POST':
        # 1. Pega a lista dos nomes "errados" que o admin marcou
        nomes_originais = request.POST.getlist('nomes_originais')
        
        # 2. Pega o nome "certo" que ele digitou
        nome_correto = request.POST.get('nome_correto')
        
        if nomes_originais and nome_correto:
            # 3. O MÁGICO UPDATE EM MASSA
            # Busca todos os itens com os nomes errados e troca pelo certo de uma vez
            registros_afetados = ItemRudimento.objects.filter(descricao__in=nomes_originais).update(descricao=nome_correto)
            
            messages.success(request, f"Sucesso! {registros_afetados} exercícios foram unificados para '{nome_correto}'.")
            return redirect('scheduler:normalizar_rudimentos')
        else:
            messages.warning(request, "Selecione pelo menos um item e digite o nome correto.")

    stats_rudimentos = (
        ItemRudimento.objects
        .values('descricao')
        .annotate(total=Count('id'))
        .order_by('descricao')
    )

    lista_completa = []

    # 2. Para cada nome diferente, buscamos exemplos reais
    for item in stats_rudimentos:
        nome_rudimento = item['descricao']
        
        # Busca as 5 últimas vezes que isso apareceu
        exemplos = (
            ItemRudimento.objects
            .filter(descricao=nome_rudimento)
            .select_related('relatorio__aula', 'relatorio__professor_que_validou')
            .prefetch_related('relatorio__aula__alunos')
            .order_by('-relatorio__aula__data_hora')[:5]
        )
        
        lista_completa.append({
            'descricao': nome_rudimento,
            'total': item['total'],
            'exemplos': exemplos
        })

    contexto = {
        'titulo': 'Normalização de Rudimentos',
        'rudimentos': lista_completa
    }
    
    return render(request, 'scheduler/admin_normalizar_rudimentos.html', contexto)


@login_required
def api_autocomplete_exercicios(request):
    term = request.GET.get('term', '')
    tipo = request.GET.get('type', 'rudimento')

    if len(term) < 2:
        return JsonResponse([], safe=False)

    results = []

    if tipo == 'rudimento':
        results = ItemRudimento.objects.filter(descricao__icontains=term)\
            .values_list('descricao', flat=True).distinct()[:10]

    return JsonResponse(list(results), safe=False)
